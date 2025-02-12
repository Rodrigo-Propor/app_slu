import openpyxl
from openpyxl.styles import numbers
import os
from datetime import datetime

# Lista de arquivos Excel no diretório
excel_files = [arquivo for arquivo in os.listdir() if arquivo.endswith('.xlsx')]


def parse_date(value):
    """Converte strings no formato 'dd-mmm-aa' para datetime"""
    months = {
        "jan": "Jan", "fev": "Feb", "mar": "Mar", "abr": "Apr", "mai": "May", "jun": "Jun",
        "jul": "Jul", "ago": "Aug", "set": "Sep", "out": "Oct", "nov": "Nov", "dez": "Dec"
    }
    try:
        if isinstance(value, str):
            for pt, en in months.items():
                value = value.replace(f"-{pt}-", f"-{en}-")
            return datetime.strptime(value, "%d-%b-%y")
    except ValueError:
        return value  # Retorna o valor original se a conversão falhar
    return value


def parse_number(value, column):
    """Converte strings para números, ajustando escala quando necessário"""
    try:
        if isinstance(value, str):
            # Remove separadores de milhar e converte para float
            value = value.replace('.', '').replace(',', '.')
            number = float(value)

            # Ajusta a escala conforme a coluna
            if column == 'B' and number > 999:  # Elevação: máx. 000,00
                return number / 1000
            elif column == 'C' and number > 999999:  # Coordenada Leste: máx. 000000,000
                return number / 1000
            elif column == 'D' and number > 9999999:  # Coordenada Norte: máx. 0000000,000
                return number / 1000
            return number
    except ValueError:
        return value  # Retorna o valor original se a conversão falhar
    return value


def find_last_row(sheet, column='A'):
    """Encontra a última linha com dados, verificando se as próximas 5 linhas estão vazias"""
    max_row = sheet.max_row

    for row in range(max_row, 0, -1):
        cell_value = sheet[f'{column}{row}'].value
        if cell_value is not None:
            # Verifica se as próximas 5 linhas estão vazias
            next_5_empty = all(
                sheet[f'{column}{row + i}'].value is None
                for i in range(1, 6)
            )
            if next_5_empty:
                return row
    return 1


def format_cells(workbook, sheet):
    """Formata as células das últimas 10 linhas com dados"""
    last_row = find_last_row(sheet)
    start_row = max(1, last_row - 9)  # Começa 10 linhas antes da última

    for row in range(start_row, last_row + 1):
        # Formata coluna A (data)
        cell_a = sheet[f'A{row}']
        if isinstance(cell_a.value, str):
            parsed_date = parse_date(cell_a.value)
            if isinstance(parsed_date, datetime):
                cell_a.value = parsed_date
        if isinstance(cell_a.value, datetime):
            cell_a.number_format = 'd-mmm-yy'

        # Formata colunas B, C e D (números decimais)
        for col in ['B', 'C', 'D']:
            cell = sheet[f'{col}{row}']
            if isinstance(cell.value, str):
                parsed_number = parse_number(cell.value, col)
                if isinstance(parsed_number, (int, float)):
                    cell.value = parsed_number
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.000'  # Sem separador de milhar e 3 casas decimais


def process_excel_files():
    """Processa todos os arquivos Excel"""
    for file_name in excel_files:
        if not os.path.exists(file_name):
            print(f"Arquivo não encontrado: {file_name}")
            continue

        try:
            workbook = openpyxl.load_workbook(file_name)

            for sheet_name in workbook.sheetnames:
                print(f"Processando: {file_name} - Aba: {sheet_name}")
                sheet = workbook[sheet_name]
                format_cells(workbook, sheet)

            workbook.save(file_name)
            print(f"Arquivo processado: {file_name}")

        except Exception as e:
            print(f"Erro ao processar {file_name}: {e}")


if __name__ == "__main__":
    process_excel_files()
