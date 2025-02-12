import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import defaultdict

# Configuração do logging
log_filename = f'excel_processing_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def get_cell_format_signature(cell):
    """
    Extrai a assinatura de formatação de uma célula, incluindo fonte, cor, borda e preenchimento.
    """
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    
    format_signature = {
        'font': {
            'name': font.name,
            'size': font.size,
            'bold': font.bold,
            'italic': font.italic,
            'color': font.color.rgb if font.color else None
        },
        'fill': {
            'fill_type': fill.fill_type,
            'start_color': fill.start_color.rgb if fill.start_color else None,
            'end_color': fill.end_color.rgb if fill.end_color else None
        },
        'border': {
            'left': border.left.style if border.left else None,
            'right': border.right.style if border.right else None,
            'top': border.top.style if border.top else None,
            'bottom': border.bottom.style if border.bottom else None
        },
        'alignment': {
            'horizontal': alignment.horizontal,
            'vertical': alignment.vertical,
            'wrap_text': alignment.wrap_text
        }
    }
    
    return str(format_signature)  # Convertendo para string para usar como chave no dicionário

def find_most_common_format(worksheet, column, start_row=3, end_row=15):
    """
    Encontra o formato mais comum em uma coluna específica entre as linhas especificadas.
    """
    format_counts = defaultdict(int)
    format_examples = {}
    
    for row in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
        cell = worksheet[f'{column}{row}']
        format_signature = get_cell_format_signature(cell)
        format_counts[format_signature] += 1
        format_examples[format_signature] = cell
    
    # Encontra o formato mais comum
    if format_counts:
        most_common_format = max(format_counts.items(), key=lambda x: x[1])
        return format_examples[most_common_format[0]], format_counts[most_common_format[0]], len(format_counts)
    return None, 0, 0

def apply_format(target_cell, source_cell):
    """
    Aplica a formatação de uma célula fonte para uma célula alvo.
    """
    if source_cell is None or target_cell is None:
        return
    
    # Copia fonte
    target_cell.font = Font(
        name=source_cell.font.name,
        size=source_cell.font.size,
        bold=source_cell.font.bold,
        italic=source_cell.font.italic,
        color=source_cell.font.color
    )
    
    # Copia preenchimento
    target_cell.fill = PatternFill(
        fill_type=source_cell.fill.fill_type,
        start_color=source_cell.fill.start_color,
        end_color=source_cell.fill.end_color
    )
    
    # Copia bordas
    target_cell.border = Border(
        left=Side(style=source_cell.border.left.style) if source_cell.border.left else None,
        right=Side(style=source_cell.border.right.style) if source_cell.border.right else None,
        top=Side(style=source_cell.border.top.style) if source_cell.border.top else None,
        bottom=Side(style=source_cell.border.bottom.style) if source_cell.border.bottom else None
    )
    
    # Copia alinhamento
    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        wrap_text=source_cell.alignment.wrap_text
    )

def process_excel_file(filepath):
    """
    Processa um arquivo Excel, aplicando formatações nas 50 células vazias após a última linha com dados.
    """
    try:
        workbook = load_workbook(filepath)
        logging.info(f"Processando arquivo: {filepath}")
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            logging.info(f"\nProcessando aba: {sheet_name}")
            
            # Para cada coluna de A até J
            for col_letter in 'ABCDEFGHIJ':
                # Encontra o formato mais comum
                template_cell, count, total_formats = find_most_common_format(worksheet, col_letter)
                
                if template_cell:
                    logging.info(f"\nColuna {col_letter}:")
                    logging.info(f"- Formato mais comum encontrado {count} vezes entre {total_formats} formatos diferentes")
                    
                    # Encontra a última linha com dados
                    last_row = 1
                    for row in range(1, worksheet.max_row + 1):
                        if worksheet[f'{col_letter}{row}'].value is not None:
                            last_row = row
                    
                    # Aplica formatação apenas nas próximas 50 células vazias
                    cells_formatted = 0
                    current_row = last_row + 1
                    
                    while cells_formatted < 50:
                        cell = worksheet[f'{col_letter}{current_row}']
                        if cell.value is None:
                            apply_format(cell, template_cell)
                            cells_formatted += 1
                        current_row += 1
                        
                        # Evita processar células além do limite razoável
                        if current_row > last_row + 100:  # Limite de segurança
                            break
                    
                    logging.info(f"- {cells_formatted} células formatadas após a linha {last_row}")
                else:
                    logging.warning(f"Nenhum formato encontrado para coluna {col_letter}")
        
        # Salva o arquivo
        workbook.save(filepath)
        logging.info(f"\nArquivo salvo com sucesso: {filepath}")
        
    except Exception as e:
        logging.error(f"Erro ao processar arquivo {filepath}: {str(e)}")
        raise

def main():
    """
    Função principal que processa todos os arquivos Excel no diretório especificado.
    """
    directory = 'media/planilhas_SLU'
    
    try:
        # Verifica se o diretório existe
        if not os.path.exists(directory):
            os.makedirs(directory)
            logging.info(f"Diretório criado: {directory}")
        
        # Lista todos os arquivos Excel no diretório
        excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
        
        if not excel_files:
            logging.warning(f"Nenhum arquivo Excel encontrado em {directory}")
            return
        
        # Processa cada arquivo
        for filename in excel_files:
            filepath = os.path.join(directory, filename)
            logging.info(f"\n{'='*50}")
            logging.info(f"Iniciando processamento do arquivo: {filename}")
            
            process_excel_file(filepath)
            
            logging.info(f"Concluído processamento do arquivo: {filename}")
            logging.info('='*50)
        
    except Exception as e:
        logging.error(f"Erro durante a execução: {str(e)}")
        raise

if __name__ == "__main__":
    logging.info("Iniciando processamento de arquivos Excel")
    main()
    logging.info("Processamento concluído")