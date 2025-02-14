import openpyxl
from openpyxl.styles import numbers
import os
from datetime import datetime
import streamlit as st

# Lista de arquivos Excel no diretório
excel_files = [arquivo for arquivo in os.listdir() if arquivo.endswith('.xlsx')]

def parse_date(value):
    """Converte strings no formato 'dd-mmm-aa' para datetime"""
    months = {
        "jan": "Jan", "fev": "Feb", "mar": "Mar", "abr": "Apr", "mai": "May", "jun": "Jun",
        "jul": "Jul", "ago": "Aug", "set": "Sep", "out": "Oct", "nov": "Nov", "dez": "Dec"
    }
    try:
        if isinstance(value, str):
            for pt, en in months.items():
                value = value.replace(f"-{pt}-", f"-{en}-")
            return datetime.strptime(value, "%d-%b-%y")
    except ValueError:
        return value  # Retorna o valor original se a conversão falhar
    return value

def parse_number(value, column):
    """Converte strings para números, ajustando escala quando necessário"""
    try:
        if isinstance(value, str):
            # Remove separadores de milhar e converte para float
            value = value.replace('.', '').replace(',', '.')
            number = float(value)

            # Ajusta a escala conforme a coluna
            if column == 'B' and number > 999:  # Elevação: máx. 000,00
                return number / 1000
            elif column == 'C' and number > 999999:  # Coordenada Leste: máx. 000000,000
                return number / 1000
            elif column == 'D' and number > 9999999:  # Coordenada Norte: máx. 0000000,000
                return number / 1000
            return number
    except ValueError:
        return value  # Retorna o valor original se a conversão falhar
    return value

def find_last_row(sheet, column='A'):
    """Encontra a última linha com dados, verificando se as próximas 5 linhas estão vazias"""
    max_row = sheet.max_row

    for row in range(max_row, 0, -1):
        cell_value = sheet[f'{column}{row}'].value
        if cell_value is not None:
            # Verifica se as próximas 5 linhas estão vazias
            next_5_empty = all(
                sheet[f'{column}{row + i}'].value is None
                for i in range(1, 6)
            )
            if next_5_empty:
                return row
    return 1

def format_cells(workbook, sheet):
    """Formata as células das últimas 20 linhas com dados"""
    last_row = find_last_row(sheet)
    st.write(f"Última linha com dados: {last_row}")

    for row in range(last_row - 19, last_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            st.write(f"Célula ({row}, {col}): {cell_value}")

    if st.button("Formatar Dados"):
        for row in range(last_row - 19, last_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if isinstance(cell.value, str):
                    cell.value = parse_date(cell.value) or parse_number(cell.value, chr(col + 64))
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.000'

        workbook.save(file_name)
        st.write(f"Arquivo processado: {file_name}")

def process_excel_files():
    """Processa todos os arquivos Excel"""
    for file_name in excel_files:
        if not os.path.exists(file_name):
            st.error(f"Arquivo não encontrado: {file_name}")
            continue

        try:
            workbook = openpyxl.load_workbook(file_name)

            for sheet_name in workbook.sheetnames:
                st.write(f"Processando: {file_name} - Aba: {sheet_name}")
                sheet = workbook[sheet_name]
                format_cells(workbook, sheet)

        except Exception as e:
            st.error(f"Erro ao processar {file_name}: {e}")

if __name__ == "__main__":
    st.title("Corrige a Formatação da Planilha")
    st.write("Selecione um arquivo Excel para processar:")
    
    file_name = st.selectbox("Arquivos Excel", excel_files)
    if file_name:
        workbook = openpyxl.load_workbook(file_name)
        sheet_names = workbook.sheetnames
        selected_sheet = st.selectbox("Selecione uma aba", sheet_names)
        
        if selected_sheet:
            sheet = workbook[selected_sheet]
            process_excel_files()
