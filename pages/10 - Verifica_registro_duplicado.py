import os
from openpyxl import load_workbook

# Diretório dos arquivos Excel
folder_path = r"D:\OneDrive\Documentos\APP\app_slu\media\planilhas_SLU"

# Verifica se o diretório existe
if not os.path.exists(folder_path):
    print(f"O diretório {folder_path} não existe.")
else:
    # Percorre todos os arquivos .xlsx na pasta
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            print(f"Processando arquivo: {file_path}")
            
            # Carrega o workbook sem alterar outras configurações
            wb = load_workbook(file_path)
            
            # Verifica se o workbook contém planilhas
            if not wb.sheetnames:
                print(f"O arquivo {filename} não contém planilhas.")
                continue
            
            # Itera por cada planilha
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"  Processando planilha: {sheet_name}")
                
                # Identifica a última linha com informação na coluna A
                last_row = None
                for row in range(ws.max_row, 0, -1):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        last_row = row
                        break
                
                if last_row is None:
                    print("    Nenhuma informação encontrada na coluna A.")
                    continue
                
                # Define o intervalo: 5 linhas acima e 5 linhas abaixo (dentro dos limites da planilha)
                start_row = max(1, last_row - 5)
                end_row = min(ws.max_row, last_row + 5)
                
                # Dicionário para registrar linhas únicas com base nos valores das 10 primeiras colunas
                seen = {}
                # Lista para armazenar os números das linhas duplicadas a serem removidas
                rows_to_delete = []
                
                # Itera sobre o intervalo definido
                for row in range(start_row, end_row + 1):
                    # Coleta os valores das colunas 1 a 10 (A a J)
                    values = tuple(ws.cell(row=row, column=col).value for col in range(1, 11))
                    if values in seen:
                        rows_to_delete.append(row)
                    else:
                        seen[values] = row
                
                if rows_to_delete:
                    print(f"    Linhas duplicadas encontradas: {rows_to_delete}")
                    # Remove as linhas duplicadas em ordem decrescente para não alterar os índices ainda não processados
                    for row in sorted(rows_to_delete, reverse=True):
                        ws.delete_rows(row, 1)
                    print("    Duplicatas removidas.")
                else:
                    print("    Nenhuma duplicata encontrada no bloco de análise.")
            
            try:
                # Salva o workbook com as alterações (somente remoção de linhas duplicadas)
                wb.save(file_path)
                print(f"Arquivo '{filename}' processado.\n")
            except OSError as e:
                print(f"Erro ao salvar o arquivo '{filename}': {e}")
