import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
from docx import Document
from docx.shared import Inches
import logging
from pathlib import Path
import excel2img
from copy import deepcopy, copy
import matplotlib.pyplot as plt
import seaborn as sns

# Configuração inicial do Streamlit
st.set_page_config(
    page_title="Processador de Planilhas de Inclinômetro",
    page_icon="📊",
    layout="wide"
)

# Criação do diretório de logs se não existir
log_dir = Path('logs')
log_dir.mkdir(parents=True, exist_ok=True)

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_dir / 'processamento.log', mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class ProcessadorPlanilhas:
    def __init__(self):
        self.base_dir = Path('.').resolve()
        self.planilha_fonte = self.base_dir / 'media' / 'planilhas_SLU' / 'Inclinômetro.xlsx'
        self.planilha_template = self.base_dir / 'media' / 'template' / 'planilha_inclinometro.xlsx'
        self.dir_temp = self.base_dir / 'temporario'
        self.abas = ["Coordenada NORTE", "Coordenada ESTE", "Cota"]

        self.criar_diretorios()

    def criar_diretorios(self):
        for dir_path in [
            self.base_dir / 'media' / 'planilhas_SLU',
            self.base_dir / 'media' / 'template',
            self.dir_temp
        ]:
            dir_path.mkdir(parents=True, exist_ok=True)
            logging.info(f"Diretório criado: {dir_path}")


    def verificar_arquivos(self):
        """Verifica se todos os arquivos necessários existem"""
        if not self.planilha_fonte.exists():
            raise FileNotFoundError(f"Planilha fonte não encontrada: {self.planilha_fonte}")
        if not self.planilha_template.exists():
            raise FileNotFoundError(f"Planilha template não encontrada: {self.planilha_template}")

    def encontrar_coluna_data(self, df):
        """Encontra a coluna de data no DataFrame"""
        for col in df.columns:
            if isinstance(col, str) and 'data' in col.lower():
                return col
        return None

    def processar_datas(self, df):
        """Processa e filtra as datas do DataFrame"""
        try:
            df = df.copy()
            coluna_data = self.encontrar_coluna_data(df)
            if not coluna_data:
                raise ValueError("Coluna de data não encontrada")
            df = df.rename(columns={coluna_data: 'Data'})
            df['Data'] = pd.to_datetime(df['Data'], format='%d/%m/%Y', errors='coerce')
            df = df.dropna(subset=['Data'])
            if df.empty:
                raise ValueError("Não há datas válidas na planilha")
            data_mais_recente = df['Data'].max()
            data_12_meses_atras = data_mais_recente - timedelta(days=365)
            logging.info(f"Intervalo de datas: {data_12_meses_atras.strftime('%d/%m/%Y')} até {data_mais_recente.strftime('%d/%m/%Y')}")
            df = df[
                (df['Data'] >= data_12_meses_atras) &
                (df['Data'] <= data_mais_recente)
            ].copy()
            df['Data'] = df['Data'].dt.strftime('%d/%m/%y')
            return df
        except Exception as e:
            logging.error(f"Erro ao processar datas: {str(e)}")
            raise

    def copiar_estilo_celula(self, cell_origem, cell_destino):
        """Copia o estilo completo de uma célula para outra"""
        if cell_origem.has_style:
            cell_destino.font = copy(cell_origem.font)
            cell_destino.border = copy(cell_origem.border)
            cell_destino.fill = copy(cell_origem.fill)
            cell_destino.alignment = copy(cell_origem.alignment)
            cell_destino.number_format = cell_origem.number_format
            cell_destino.protection = copy(cell_origem.protection)

    def salvar_tabela_como_imagem_alternativo(self, ws_template, nome_arquivo_imagem, df):
        """Método alternativo para salvar a tabela como imagem usando matplotlib"""
        try:
            fig, ax = plt.subplots(figsize=(12, len(df) * 0.4))
            ax.axis('off')
            table = ax.table(
                cellText=df.values,
                colLabels=df.columns,
                cellLoc='center',
                loc='center',
                colWidths=[0.1] * len(df.columns)
            )
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1, 1.5)
            plt.savefig(
                nome_arquivo_imagem,
                bbox_inches='tight',
                dpi=300,
                pad_inches=0.1
            )
            plt.close()
            logging.info(f"Imagem salva com método alternativo: {nome_arquivo_imagem}")
        except Exception as e:
            logging.error(f"Erro ao salvar imagem com método alternativo: {str(e)}")
            raise

    def salvar_tabela_como_imagem(self, ws_template, nome_arquivo_imagem):
        """Salva a planilha como imagem tentando múltiplos métodos"""
        try:
            temp_excel = self.dir_temp / 'temp_for_image.xlsx'
            try:
                if temp_excel.exists():
                    os.remove(temp_excel)
                wb = ws_template.parent
                wb.save(str(temp_excel.absolute()))
                excel2img.export_img(
                    str(temp_excel.absolute()),
                    str(Path(nome_arquivo_imagem).absolute()),
                    ws_template.title,
                    'A1:J31'
                )
                if temp_excel.exists():
                    os.remove(temp_excel)
                logging.info(f"Imagem salva com sucesso usando excel2img: {nome_arquivo_imagem}")
            except Exception as e1:
                logging.warning(f"Falha ao usar excel2img, tentando método alternativo: {str(e1)}")
                if temp_excel.exists():
                    os.remove(temp_excel)
                data = []
                for row in ws_template.iter_rows(min_row=1, values_only=True):
                    if any(cell is not None for cell in row):
                        data.append(row)
                df = pd.DataFrame(data[1:], columns=data[0])
                self.salvar_tabela_como_imagem_alternativo(ws_template, nome_arquivo_imagem, df)
        except Exception as e:
            logging.error(f"Erro ao salvar imagem (ambos os métodos falharam): {str(e)}")
            raise

    def inserir_imagem_no_word(self, caminho_imagem, arquivo_word):
        """Insere a imagem no documento Word com ajustes de qualidade"""
        try:
            doc = Document()
            doc.add_picture(caminho_imagem, width=Inches(6))
            doc.save(arquivo_word)
            logging.info(f"Documento Word criado com sucesso: {arquivo_word}")
        except Exception as e:
            logging.error(f"Erro ao criar documento Word: {str(e)}")
            raise

    def processar_aba(self, nome_aba, progress_bar):
        """Processa uma aba específica da planilha"""
        try:
            logging.info(f"Iniciando processamento da aba: {nome_aba}")
            wb_template = openpyxl.load_workbook(self.planilha_template)
            ws_template = wb_template[nome_aba]
            estilos_template = {}
            for row in range(1, 31):
                for col in range(1, 11):
                    cell = ws_template.cell(row=row, column=col)
                    estilos_template[(row, col)] = cell
            progress_bar.progress(0.2)
            df_fonte = pd.read_excel(
                self.planilha_fonte,
                sheet_name=nome_aba,
                header=None
            )
            header_row = None
            for idx, row in df_fonte.iterrows():
                if row.astype(str).str.contains('Data', case=False).any():
                    header_row = idx
                    df_fonte = pd.read_excel(
                        self.planilha_fonte,
                        sheet_name=nome_aba,
                        header=header_row
                    )
                    break
            if header_row is None:
                raise ValueError("Cabeçalho não encontrado na planilha fonte")
            progress_bar.progress(0.4)
            df_processado = self.processar_datas(df_fonte)
            template_columns = [cell.value for cell in ws_template[1] if cell.value]

            # Adicionar colunas faltantes ao df_processado
            for col in template_columns:
                if col not in df_processado.columns:
                    df_processado[col] = ""  # Ou pd.NA, dependendo do que você prefere

            df_processado = df_processado[template_columns] #Reordenar colunas

            progress_bar.progress(0.6)
            for row in range(2, 31):
                for col in range(1, 11):
                    ws_template.cell(row=row, column=col).value = None
            for i, row in enumerate(dataframe_to_rows(df_processado, index=False, header=False), 2):
                for j, value in enumerate(row, 1):
                    cell_destino = ws_template.cell(row=i, column=j)
                    cell_destino.value = value
                    cell_template = estilos_template.get((2, j))
                    if cell_template:
                        self.copiar_estilo_celula(cell_template, cell_destino)
            progress_bar.progress(0.8)
            temp_excel = self.dir_temp / f'temp_{nome_aba.replace(" ", "_")}.xlsx'
            wb_template.save(temp_excel)
            nome_arquivo_imagem = self.dir_temp / f'tabela_{nome_aba.replace(" ", "_")}.png'
            self.salvar_tabela_como_imagem(ws_template, str(nome_arquivo_imagem))
            nome_arquivo = f'inclinometro_{nome_aba.split()[-1].lower()}.docx'
            arquivo_word = self.dir_temp / nome_arquivo
            self.inserir_imagem_no_word(str(nome_arquivo_imagem), arquivo_word)
            if nome_arquivo_imagem.exists():
                os.remove(str(nome_arquivo_imagem))
            if temp_excel.exists():
                os.remove(temp_excel)
            progress_bar.progress(1.0)
            logging.info(f"Aba {nome_aba} processada com sucesso")
            return True, "Processamento concluído com sucesso!"
        except Exception as e:
            logging.error(f"Erro ao processar aba {nome_aba}: {str(e)}")
            return False, f"Erro ao processar aba {nome_aba}: {str(e)}"


def main():
    st.title("Processador de Planilhas de Inclinômetro")
    st.write("Este aplicativo processa dados de inclinômetro e gera relatórios em Word.")
    processador = ProcessadorPlanilhas()
    st.sidebar.title("Status dos Arquivos")
    try:
        processador.verificar_arquivos()
        st.sidebar.success("✅ Diretórios e arquivos verificados")
        st.sidebar.success(f"📁 Fonte: {processador.planilha_fonte.name}")
        st.sidebar.success(f"📁 Template: {processador.planilha_template.name}")
    except Exception as e:
        st.sidebar.error(f"❌ Erro: {str(e)}")
        return
    if st.button("Iniciar Processamento"):
        overall_progress = st.progress(0)
        try:
            for idx, aba in enumerate(processador.abas):
                st.write(f"Processando aba: {aba}")
                progress_bar = st.progress(0)
                sucesso, mensagem = processador.processar_aba(aba, progress_bar)
                if sucesso:
                    st.success(f"✅ Aba {aba} processada com sucesso!")
                else:
                    st.error(f"❌ Erro ao processar aba {aba}: {mensagem}")
                    break
                overall_progress.progress((idx + 1) / len(processador.abas))
            if all(processador.dir_temp.joinpath(f'inclinometro_{aba.split()[-1].lower()}.docx').exists()
                   for aba in processador.abas):
                st.success("🎉 Processamento completo!")
                st.write("### 📥 Downloads Disponíveis")
                cols = st.columns(len(processador.abas))
                for aba, col in zip(processador.abas, cols):
                    nome_arquivo = f'inclinometro_{aba.split()[-1].lower()}.docx'
                    arquivo_word = processador.dir_temp / nome_arquivo
                    if arquivo_word.exists():
                        with open(arquivo_word, "rb") as f:
                            col.download_button(
                                label=f"📄 Baixar {aba}",
                                data=f.read(),
                                file_name=nome_arquivo,
                                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                                use_container_width=True
                            )
        except Exception as e:
            st.error(f"❌ Erro durante o processamento: {str(e)}")
            logging.error(f"Erro durante o processamento: {str(e)}")
        if (log_dir / 'processamento.log').exists():
            with st.expander("📋 Ver Logs"):
                with open(log_dir / 'processamento.log', 'r', encoding='utf-8') as f:
                    st.code(f.read())

if __name__ == "__main__":
    main()
