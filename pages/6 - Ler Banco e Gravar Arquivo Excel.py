import os
import sqlite3
import pandas as pd
from datetime import datetime, date
import re
import zipfile
import tkinter as tk
from tkinter import messagebox
import time
import streamlit as st
from openpyxl import load_workbook, Workbook

# Diretórios e configuração de arquivos
media_path = 'media/planilhas_SLU'
db_path = 'banco_dados.db'
arquivos_db_path = 'arquivos.db'
required_files = [
    "Placa de Recalque AC 01.xlsx", "Placa de Recalque AC 03.xlsx",
    "Placa de Recalque AC 04.xlsx", "Placa de Recalque AC 05.xlsx",
    "Placa de Recalque AMPLIAÇÃO.xlsx", "Placa de Recalque Emergencial.xlsx",
    "Recalques Celula Pesquisa.xlsx"
]
# Incluir o arquivo inclinômetro
inclinometro_file = "Inclinômetro.xlsx"

contador_processados = 0
contador_duplicados = 0
contador_erros = 0

# ------------------------- FUNÇÕES AUXILIARES -------------------------
def formatar_numero_br(valor):
    return f"{valor:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")

def criar_tabela_arquivos(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ultimas_datas (
            arquivo TEXT,
            aba TEXT,
            ultima_data TEXT,
            PRIMARY KEY(arquivo, aba)
        )
    """)
    conn.commit()

def atualizar_datas_arquivos():
    """
    Para cada arquivo (dos required_files e inclinômetro), abre com openpyxl,
    lê a última data da coluna A (assumindo formato 'dd-mmm-yy') de cada aba e
    grava/atualiza essa informação na tabela ultimas_datas do banco arquivos.db.
    """
    conn = sqlite3.connect(arquivos_db_path)
    criar_tabela_arquivos(conn)
    arquivos_para_verificar = required_files + [inclinometro_file]
    for arq in arquivos_para_verificar:
        file_path = os.path.join(media_path, arq)
        if not os.path.exists(file_path):
            continue
        wb = load_workbook(file_path, data_only=True)
        for aba in wb.sheetnames:
            ws = wb[aba]
            ultima_data = None
            # Assume que a coluna A (linha 1 em diante) possui datas
            for row in range(1, ws.max_row+1):
                valor = ws.cell(row=row, column=1).value
                if valor:
                    try:
                        # Se for string, converter; se for data, usar diretamente
                        if isinstance(valor, str):
                            data_valor = datetime.strptime(valor.strip(), "%d-%b-%y")
                        elif isinstance(valor, (datetime, date)):
                            data_valor = valor if isinstance(valor, datetime) else datetime.combine(valor, datetime.min.time())
                        else:
                            continue
                        ultima_data = data_valor
                    except Exception:
                        continue
            if ultima_data:
                # Armazena a última data em formato ISO (YYYY-MM-DD)
                conn.execute("""
                    INSERT OR REPLACE INTO ultimas_datas (arquivo, aba, ultima_data)
                    VALUES (?, ?, ?)
                """, (arq, aba, ultima_data.strftime("%Y-%m-%d")))
    conn.commit()
    conn.close()

def obter_ultima_data(arquivo, aba):
    """
    Retorna a última data registrada (como datetime) para o arquivo e aba
    da tabela ultimas_datas. Se não existir, retorna None.
    """
    conn = sqlite3.connect(arquivos_db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT ultima_data FROM ultimas_datas WHERE arquivo=? AND aba=?", (arquivo, aba))
    row = cursor.fetchone()
    conn.close()
    if row and row[0]:
        return datetime.strptime(row[0], "%Y-%m-%d")
    return None

def escrever_registros_no_excel(file_path, aba_nome, registros):
    """
    Abre o arquivo Excel, seleciona ou cria a aba 'aba_nome' e escreve
    os registros (lista de dicionários com chaves "Data", "Elevação",
    "Coordenada Este", "Coordenada Norte").
    Os dados serão adicionados abaixo da última linha preenchida (a partir da linha 3).
    """
    wb = load_workbook(file_path)
    if aba_nome in wb.sheetnames:
        ws = wb[aba_nome]
    else:
        ws = wb.create_sheet(aba_nome)
    linha = 3
    while ws.cell(row=linha, column=1).value is not None:
        linha += 1
    for dado in registros:
        ws[f"A{linha}"] = dado["Data"]
        ws[f"B{linha}"] = dado["Elevação"]
        ws[f"C{linha}"] = dado["Coordenada Este"]
        ws[f"D{linha}"] = dado["Coordenada Norte"]
        linha += 1
    wb.save(file_path)

# Para os registros de Inclinômetro, definindo mapeamento de pontos
# Observação: as colunas C a J correspondem a: I8, I7, I3, I2, I1, I6, I4, I5
colunas_mapping = {
    "I8": "C",
    "I7": "D",
    "I3": "E",
    "I2": "F",
    "I1": "G",
    "I6": "H",
    "I4": "I",
    "I5": "J"
}
# Função para calcular número de dias decorridos desde 13/04/2016
def dias_desde_base(data_reg):
    base = datetime.strptime("13/04/2016", "%d/%m/%Y")
    return (data_reg - base).days

def escrever_inclinometro(file_path, dados_inclinometro):
    """
    dados_inclinometro é um dicionário com chaves: "Coordenada NORTE", "Coordenada ESTE" e "Cota".
    Cada um é outro dicionário agrupado por data (em formato datetime) contendo:
       { data: { "I1": valor, "I2": valor, ... } }
    Para cada aba, abre ou cria a planilha, procura se já existe linha para a data (com valor em coluna A);
    se existir, atualiza as células correspondentes; senão, insere nova linha com:
       Coluna A: data (formatada como "dd-mmm-yy")
       Coluna B: número de dias decorridos
       Colunas C a J: valores conforme mapeamento.
    """
    wb = load_workbook(file_path)
    for aba_nome, grupo in dados_inclinometro.items():
        if aba_nome in wb.sheetnames:
            ws = wb[aba_nome]
        else:
            ws = wb.create_sheet(aba_nome)
        # Processa cada data (ordenada)
        for data_reg, pontos in sorted(grupo.items()):
            data_str = data_reg.strftime("%d-%b-%y")
            # Procura se a data já existe na coluna A
            linha = None
            for r in range(1, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == data_str:
                    linha = r
                    break
            if not linha:
                # Se não existir, insere nova linha no final
                linha = ws.max_row + 1
                ws[f"A{linha}"] = data_str
                ws[f"B{linha}"] = dias_desde_base(data_reg)
            # Atualiza a célula correspondente conforme placa
            for placa, valor in pontos.items():
                coluna = colunas_mapping.get(placa)
                if coluna:
                    ws[f"{coluna}{linha}"] = valor
    wb.save(file_path)

# ------------------------- FUNÇÕES PRINCIPAIS -------------------------
def processar_registros():
    global contador_processados, contador_duplicados, contador_erros

    # Antes de processar os registros, atualiza o banco de dados de última data por arquivo/aba
    atualizar_datas_arquivos()

    # Dicionários para agrupar registros a inserir
    registros_por_grupo = {}  # chave: (arquivo_destino, aba) -> lista de dict com dados
    inclinometro_dados = { "Coordenada NORTE": {}, "Coordenada ESTE": {}, "Cota": {} }  # agrupados por data

    # Conectar ao banco de dados dos registros originais
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM placas_completas_slu_bh")
    registros = cursor.fetchall()
    conn.close()

    for registro in registros:
        # Campos: id, tipo, descricao, coordenada_este, coordenada_norte, elevacao,
        # info_gerada, nome_arquivo, placa, data
        id, tipo, descricao, coord_este, coord_norte, elevacao, info_gerada, nome_arquivo, placa, data_str = registro
        try:
            data_registro = datetime.strptime(data_str, "%Y-%m-%d")
        except Exception as e:
            print(f"Erro na conversão da data do registro {id}: {e}")
            contador_erros += 1
            continue

        # Se for registro de inclinômetro (agora incluído)
        if tipo == "INCLINOMETRO":
            # Nos registros inclinômetro o campo "placa" indica o ponto (ex: "I1", "I2", etc)
            ponto = placa.strip() if placa else ""
            if ponto not in colunas_mapping:
                print(f"Aviso: Ponto '{ponto}' não reconhecido para INCLINOMETRO.")
                continue
            # Para cada aba, a lógica é similar: cada aba recebe um valor específico
            # 'Coordenada NORTE': usa coord_norte; 'Coordenada ESTE': usa coord_este; 'Cota': usa elevacao
            for aba_nome, valor in [("Coordenada NORTE", coord_norte),
                                    ("Coordenada ESTE", coord_este),
                                    ("Cota", elevacao)]:
                grupo = inclinometro_dados[aba_nome]
                # Agrupa por data: se já existir, atualiza o dicionário interno
                if data_registro not in grupo:
                    grupo[data_registro] = {}
                grupo[data_registro][ponto] = formatar_numero_br(valor)
            contador_processados += 1
            continue

        # Processa registros dos outros tipos: CELULA EMERGENCIAL, CELULA PESQUISA, PAMPULHA
        if tipo not in ("CELULA EMERGENCIAL", "CELULA PESQUISA", "PAMPULHA"):
            continue

        # Determina o arquivo destino conforme o tipo e regras para PAMPULHA
        if tipo == "CELULA EMERGENCIAL":
            arquivo_destino = "Placa de Recalque Emergencial.xlsx"
        elif tipo == "CELULA PESQUISA":
            arquivo_destino = "Recalques Celula Pesquisa.xlsx"
        elif tipo == "PAMPULHA":
            if placa and (placa.startswith("PR 1") or placa in ("D1", "D2")):
                arquivo_destino = "Placa de Recalque AC 01.xlsx"
            elif placa and placa.startswith("PR 3"):
                arquivo_destino = "Placa de Recalque AC 03.xlsx"
            elif placa and placa.startswith("PR 4"):
                arquivo_destino = "Placa de Recalque AC 04.xlsx"
            elif placa and placa.startswith("PR 5"):
                arquivo_destino = "Placa de Recalque AC 05.xlsx"
            elif placa and placa.startswith("PR A"):
                arquivo_destino = "Placa de Recalque AMPLIAÇÃO.xlsx"
            else:
                print(f"Aviso: Registro do tipo PAMPULHA com aba '{placa}' não corresponde a nenhum critério de destino.")
                continue

        # A aba onde será gravado é definida pelo campo placa (remove espaços)
        aba_desejada = placa.strip() if placa else ""
        # Consulta a última data já gravada para este arquivo/aba
        ultima_data = obter_ultima_data(arquivo_destino, aba_desejada)
        if ultima_data and data_registro <= ultima_data:
            # Se a data do registro já está contida, ignora
            continue

        # Prepara os dados a serem gravados
        dados = {
            "Data": data_registro.strftime("%d-%b-%y"),
            "Elevação": formatar_numero_br(elevacao),
            "Coordenada Este": formatar_numero_br(coord_este),
            "Coordenada Norte": formatar_numero_br(coord_norte)
        }
        chave = (arquivo_destino, aba_desejada)
        if chave not in registros_por_grupo:
            registros_por_grupo[chave] = []
        registros_por_grupo[chave].append(dados)
        contador_processados += 1

    # Agora, para cada grupo (arquivo, aba), escreve os registros no Excel de uma vez
    for (arq, aba), regs in registros_por_grupo.items():
        file_path = os.path.join(media_path, arq)
        escrever_registros_no_excel(file_path, aba, regs)

    # Processa os dados do inclinômetro (se houver)
    if inclinometro_dados["Coordenada NORTE"]:
        file_path_inc = os.path.join(media_path, inclinometro_file)
        escrever_inclinometro(file_path_inc, inclinometro_dados)

    # Atualiza o banco de datas após a gravação (pode-se chamar novamente atualizar_datas_arquivos())
    atualizar_datas_arquivos()

    # Exibe relatório final
    print(f"""
    Processamento concluído:
    - Registros processados: {contador_processados}
    - Registros duplicados ignorados: {contador_duplicados}
    - Erros encontrados: {contador_erros}
    """)

# ------------------------- INTERFACE STREAMLIT -------------------------
def interface_streamlit():
    st.title("Processamento de Planilhas SLU")
    tab1, tab2 = st.tabs(["Informações", "Processamento"])
    
    with tab1:
        st.markdown("""
        ### Sobre o Processamento
        Este programa realiza as seguintes operações:
        1. Verifica a última data registrada em cada arquivo Excel e atualiza o banco arquivos.db.
        2. Filtra e organiza os registros novos da tabela placas_completas_slu_bh para atualização dos arquivos.
        3. Inclui o processamento do arquivo Inclinômetro.xlsx, gravando as coordenadas organizadas por data.
        """)
        st.subheader("Arquivos Necessários")
        for arq in required_files + [inclinometro_file]:
            st.write(f"- {arq}")
    
    with tab2:
        # Verificar existência dos arquivos
        missing = [f for f in (required_files + [inclinometro_file]) if not os.path.exists(os.path.join(media_path, f))]
        if missing:
            st.error(f"Erro: Os seguintes arquivos estão ausentes na pasta '{media_path}': {', '.join(missing)}")
            return

        status_container = st.empty()
        progress_container = st.empty()
        
        col1, col2 = st.columns(2)
        with col1:
            iniciar = st.button("Iniciar Processamento", type="primary")
        with col2:
            confirmar = st.checkbox("Confirmar início do processamento")
        
        if iniciar:
            if not confirmar:
                st.warning("Por favor, confirme para iniciar o processamento")
                return
            status_container.info("Iniciando processamento...")
            try:
                processar_registros()
                status_container.success("Processamento concluído!")
            except Exception as e:
                status_container.error(f"Erro durante o processamento: {str(e)}")
                st.error("O processamento foi interrompido devido a um erro.")

if __name__ == "__main__":
    interface_streamlit()