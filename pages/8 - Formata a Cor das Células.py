import os
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from collections import defaultdict
import streamlit as st

# Configuração do logging
log_filename = f'excel_processing_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def get_cell_format_signature(cell):
    """
    Extrai a assinatura de formatação de uma célula, incluindo fonte, cor, borda e preenchimento.
    """
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    
    format_signature = {
        'font': {
            'name': font.name,
            'size': font.size,
            'bold': font.bold,
            'italic': font.italic,
            'color': font.color.rgb if font.color else None
        },
        'fill': {
            'fill_type': fill.fill_type,
            'start_color': fill.start_color.rgb if fill.start_color else None,
            'end_color': fill.end_color.rgb if fill.end_color else None
        },
        'border': {
            'left': border.left.style if border.left else None,
            'right': border.right.style if border.right else None,
            'top': border.top.style if border.top else None,
            'bottom': border.bottom.style if border.bottom else None
        },
        'alignment': {
            'horizontal': alignment.horizontal,
            'vertical': alignment.vertical,
            'wrap_text': alignment.wrap_text
        }
    }
    
    return str(format_signature)  # Convertendo para string para usar como chave no dicionário

def find_most_common_format(worksheet, column, start_row=3, end_row=15):
    """
    Encontra o formato mais comum em uma coluna específica entre as linhas especificadas.
    """
    format_counts = defaultdict(int)
    format_examples = {}
    
    for row in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
        cell = worksheet[f'{column}{row}']
        format_signature = get_cell_format_signature(cell)
        format_counts[format_signature] += 1
        format_examples[format_signature] = cell
    
    # Encontra o formato mais comum
    if format_counts:
        most_common_format = max(format_counts.items(), key=lambda x: x[1])
        return format_examples[most_common_format[0]], format_counts[most_common_format[0]], len(format_counts)
    return None, 0, 0

def apply_format(target_cell, source_cell):
    """
    Aplica a formatação de uma célula fonte para uma célula alvo.
    """
    if source_cell is None or target_cell is None:
        return
    
    # Copia fonte
    target_cell.font = Font(
        name=source_cell.font.name,
        size=source_cell.font.size,
        bold=source_cell.font.bold,
        italic=source_cell.font.italic,
        color=source_cell.font.color
    )
    
    # Copia preenchimento
    target_cell.fill = PatternFill(
        fill_type=source_cell.fill.fill_type,
        start_color=source_cell.fill.start_color,
        end_color=source_cell.fill.end_color
    )
    
    # Copia bordas
    target_cell.border = Border(
        left=Side(style=source_cell.border.left.style) if source_cell.border.left else None,
        right=Side(style=source_cell.border.right.style) if source_cell.border.right else None,
        top=Side(style=source_cell.border.top.style) if source_cell.border.top else None,
        bottom=Side(style=source_cell.border.bottom.style) if source_cell.border.bottom else None
    )
    
    # Copia alinhamento
    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        wrap_text=source_cell.alignment.wrap_text
    )

def format_column_as_integer(worksheet, column):
    """
    Formata todas as células de uma coluna específica como números inteiros com precisão de 0 casas decimais.
    """
    for row in range(1, worksheet.max_row + 1):
        cell = worksheet[f'{column}{row}']
        if cell.value is not None:
            try:
                cell.value = round(float(cell.value))
                cell.number_format = '0'
            except ValueError:
                pass  # Ignora células que não podem ser convertidas para inteiro

def process_excel_file(filepath, progress_bar=None, status_text=None):
    """
    Processa um arquivo Excel, aplicando formatações nas últimas 3 linhas com dados na coluna A e nas 50 células vazias após a última linha com dados.
    """
    try:
        workbook = load_workbook(filepath)
        if status_text:
            status_text.write(f"Processando arquivo: {filepath}")
        
        total_sheets = len(workbook.sheetnames)
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet_name]
            if status_text:
                status_text.write(f"Processando aba: {sheet_name}")
            
            # Atualiza progresso
            if progress_bar:
                progress = (sheet_idx / total_sheets)
                progress_bar.progress(progress)
            
            # Verifica se o arquivo é "Inclinômetro.xlsx" e formata a coluna B como inteiro
            if os.path.basename(filepath) == "Inclinômetro.xlsx":
                format_column_as_integer(worksheet, 'B')
            
            # Para cada coluna de A até J
            for col_letter in 'ABCDEFGHIJ':
                template_cell, count, total_formats = find_most_common_format(worksheet, col_letter)
                
                if template_cell:
                    if status_text:
                        status_text.write(f"Formatando coluna {col_letter}... (Formatos encontrados: {total_formats})")
                    
                    # Encontra a última linha com dados na coluna A
                    last_row = 1
                    for row in range(1, worksheet.max_row + 1):
                        if worksheet['A' + str(row)].value is not None:
                            last_row = row
                    
                    # Aplica formatação nas últimas 3 linhas com dados na coluna A
                    for row in range(last_row - 2, last_row + 1):
                        cell = worksheet[f'{col_letter}{row}']
                        apply_format(cell, template_cell)
                    
                    # Aplica formatação nas próximas 50 células vazias
                    cells_formatted = 0
                    current_row = last_row + 1
                    
                    while cells_formatted < 50:
                        cell = worksheet[f'{col_letter}{current_row}']
                        if cell.value is None:
                            apply_format(cell, template_cell)
                            cells_formatted += 1
                        current_row += 1
                        
                        # Evita processar células além do limite razoável
                        if current_row > last_row + 100:  # Limite de segurança
                            break
                    
                    if status_text:
                        status_text.write(f"- {cells_formatted} células formatadas após a linha {last_row}")
                else:
                    if status_text:
                        status_text.warning(f"Nenhum formato encontrado para coluna {col_letter}")
        
        # Salva o arquivo
        workbook.save(filepath)
        if status_text:
            status_text.success(f"Arquivo salvo com sucesso: {filepath}")
        
    except Exception as e:
        if status_text:
            status_text.error(f"Erro ao processar arquivo {filepath}: {str(e)}")
        raise

def main():
    st.title("Formatação de Células em Arquivos Excel")
    
    # Criando tabs para melhor organização
    tab1, tab2 = st.tabs(["Informações", "Processamento"])
    
    with tab1:
        st.write("""
        ### Sobre o Processamento
        Este programa realiza as seguintes operações:
        1. Analisa arquivos Excel na pasta `media/planilhas_SLU`
        2. Identifica o formato mais comum nas células das colunas A até J
        3. Aplica esse formato nas próximas 50 células vazias após a última linha com dados
        4. Salva as alterações mantendo a formatação original
        
        ### Arquivos Processados
        Os seguintes tipos de arquivos são processados:
        - Arquivos Excel (.xlsx) encontrados na pasta
        - Apenas as colunas de A até J são formatadas
        - São formatadas 50 células vazias após a última linha com dados
        """)
    
    with tab2:
        directory = 'media/planilhas_SLU'
        
        # Verifica se o diretório existe
        if not os.path.exists(directory):
            if st.button("Criar diretório"):
                os.makedirs(directory)
                st.success(f"Diretório criado: {directory}")
        
        # Lista arquivos Excel
        excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
        
        if not excel_files:
            st.warning(f"Nenhum arquivo Excel encontrado em {directory}")
            st.stop()
        
        st.write("### Arquivos Encontrados")
        for file in excel_files:
            st.write(f"- {file}")
        
        # Botão para iniciar processamento
        if st.button("Iniciar Processamento"):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Processa cada arquivo
                for idx, filename in enumerate(excel_files):
                    filepath = os.path.join(directory, filename)
                    status_text.write(f"Processando arquivo {idx+1} de {len(excel_files)}: {filename}")
                    
                    process_excel_file(filepath, progress_bar, status_text)
                    
                    # Atualiza barra de progresso geral
                    progress = (idx + 1) / len(excel_files)
                    progress_bar.progress(progress)
                
                st.success("Processamento concluído com sucesso!")
                
            except Exception as e:
                st.error(f"Erro durante o processamento: {str(e)}")
                logging.error(f"Erro durante a execução: {str(e)}")

if __name__ == "__main__":
    main()