import os
import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import streamlit as st  # Importar Streamlit

# Caminho para o banco de dados
CAMINHO_BANCO = r"D:\OneDrive\Documentos\APP\app_slu\banco_dados.db"

# Diretório onde estão os arquivos Excel
DIRETORIO_ARQUIVOS = r"D:\OneDrive\Documentos\APP\app_slu\media\planilhas_SLU"

# Lista de arquivos de interesse (para filtrar apenas esses, se quiser)
lista_arquivos = [
    "Inclinômetro.xlsx",
    "Placa de Recalque AC 01.xlsx",
    "Placa de Recalque AC 03.xlsx",
    "Placa de Recalque AC 04.xlsx",
    "Placa de Recalque AC 05.xlsx",
    "Placa de Recalque AMPLIAÇÃO.xlsx",
    "Placa de Recalque Emergencial.xlsx",
    "Recalques Celula Pesquisa.xlsx"
]

# Função auxiliar para verificar e formatar datas, se necessário
def formata_data(valor_celula):
    """
    Se o valor for datetime, converte para dd/mm/yyyy.
    Caso contrário, converte apenas para string.
    """
    if isinstance(valor_celula, datetime):
        return valor_celula.strftime("%d/%m/%Y")
    return str(valor_celula)

# Função para padronizar datas
def padroniza_data(valor):
    from datetime import datetime
    if isinstance(valor, datetime):
        return valor.date()
    formatos_data = ["%d/%m/%y", "%d/%m/%Y", "%d-%b-%y", "%d-%b-%Y", "%d/%b/%y", "%d/%b/%Y"]
    for formato in formatos_data:
        try:
            return datetime.strptime(str(valor), formato).date()
        except:
            pass
    return None

# Função principal para processar os arquivos
def processar_arquivos():
    # 1 - Criar conexão com o banco de dados SQLite e a tabela necessária
    con = sqlite3.connect(CAMINHO_BANCO)
    cur = con.cursor()

    # Cria a tabela se não existir
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ultima_linha_arquivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_arquivo TEXT,
            nome_planilha TEXT,
            linha_informacao INTEGER,
            data_ultimo_registro TEXT
        )
        """
    )
    con.commit()

    # Apaga todos os registros anteriores
    cur.execute("DELETE FROM ultima_linha_arquivos")
    con.commit()

    # 2 - Percorrer os arquivos Excel no diretório
    for arquivo in os.listdir(DIRETORIO_ARQUIVOS):
        if arquivo.endswith(".xlsx") and arquivo in lista_arquivos:
            caminho_arquivo = os.path.join(DIRETORIO_ARQUIVOS, arquivo)
            st.write(f"Processando arquivo: {caminho_arquivo}")

            # 3 - Abrir o arquivo Excel com openpyxl
            wb = load_workbook(caminho_arquivo, data_only=False, read_only=False)  # data_only=False e read_only=False para leitura-escrita

            # Percorrer cada planilha
            for nome_planilha in wb.sheetnames:
                ws = wb[nome_planilha]
                st.write(f"  Planilha: {nome_planilha}")

                # Encontrar a última linha com informação na coluna A a partir da linha 10
                registros_validos = []
                for row in range(ws.max_row, 9, -1):
                    valor_celula = ws.cell(row=row, column=1).value
                    if valor_celula not in (None, "") and str(valor_celula).strip() != "":
                        # Verificar se a data é de 2024 em diante
                        if isinstance(valor_celula, datetime) and valor_celula.year >= 2024:
                            # Verificar se as 7 linhas anteriores estão preenchidas
                            linhas_validas = True
                            for i in range(1, 8):
                                valor_anterior = ws.cell(row=row-i, column=1).value
                                if valor_anterior in (None, "") or str(valor_anterior).strip() == "":
                                    linhas_validas = False
                                    break
                            if not linhas_validas:
                                st.write(f"    Registro inválido encontrado na linha {row}. Apagando...")
                                for col in range(1, ws.max_column + 1):
                                    ws.cell(row=row, column=col).value = None
                                # Salvar as alterações no arquivo Excel
                                wb.save(caminho_arquivo)
                                continue
                        registros_validos.append(row)

                # Verificar duplicados por data
                from math import sqrt
                duplicados = {}
                for r in registros_validos:
                    data_crua = ws.cell(row=r, column=1).value
                    data_padronizada = padroniza_data(data_crua)
                    duplicados.setdefault(data_padronizada, []).append(r)

                for data_duplicada, rows in duplicados.items():
                    if data_duplicada and data_duplicada.year >= 2024 and len(rows) > 1:
                        # Calcular média das últimas 5 coordenadas
                        ultimos_cinco = rows[-5:] if len(rows) > 5 else rows
                        coords = []
                        for rr in ultimos_cinco:
                            x = ws.cell(row=rr, column=2).value or 0
                            y = ws.cell(row=rr, column=3).value or 0
                            z = ws.cell(row=rr, column=4).value or 0
                            coords.append((x, y, z))
                        media_x = sum(c[0] for c in coords) / len(coords)
                        media_y = sum(c[1] for c in coords) / len(coords)
                        media_z = sum(c[2] for c in coords) / len(coords)

                        # Escolher o mais próximo da média como válido
                        distancias = []
                        for rr in rows:
                            x = ws.cell(row=rr, column=2).value or 0
                            y = ws.cell(row=rr, column=3).value or 0
                            z = ws.cell(row=rr, column=4).value or 0
                            dist = sqrt((x - media_x)**2 + (y - media_y)**2 + (z - media_z)**2)
                            distancias.append((dist, rr))
                        distancias.sort(key=lambda d: d[0])
                        valido = distancias[0][1]
                        # Apagar os outros
                        for dist, rr in distancias[1:]:
                            for col in range(1, ws.max_column + 1):
                                ws.cell(row=rr, column=col).value = None

                wb.save(caminho_arquivo)

                # Retoma lógica de encontrar ultima_linha e inserir no banco
                ultima_linha = None
                for row in range(ws.max_row, 9, -1):
                    valor_celula = ws.cell(row=row, column=1).value
                    if valor_celula not in (None, "") and str(valor_celula).strip() != "":
                        ultima_linha = row
                        break

                if ultima_linha is None:
                    st.write("    Nenhuma célula preenchida na coluna A. Pulando...")
                    continue

                # Pega o valor da célula nessa última linha encontrada
                valor_data = ws.cell(row=ultima_linha, column=1).value
                data_formatada = formata_data(valor_data)

                # Gravar no banco de dados (nome do arquivo, nome da planilha, linha e data)
                cur.execute(
                    """
                    INSERT INTO ultima_linha_arquivos (nome_arquivo, nome_planilha, linha_informacao, data_ultimo_registro)
                    VALUES (?, ?, ?, ?)
                    """,
                    (arquivo, nome_planilha, ultima_linha, data_formatada)
                )
                con.commit()

                st.write(f"    >> Última linha: {ultima_linha} | Data: {data_formatada}")

                # Salvar as alterações no arquivo Excel
                wb.save(caminho_arquivo)

            st.write(f"Arquivo '{arquivo}' finalizado.\n")

            # Fechar o workbook
            wb.close()

    # Fecha a conexão com o banco ao final
    con.close()
    st.write("Processo concluído.")

# Interface do Streamlit
st.title("Processamento de Arquivos Excel")
if st.button("Iniciar Processamento"):
    processar_arquivos()
