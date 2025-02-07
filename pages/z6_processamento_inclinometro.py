import os
import sqlite3
import pandas as pd
from datetime import datetime
import logging
from openpyxl import load_workbook
import re
import locale
import traceback

# Configuração do logging
logging.basicConfig(
    filename='processo_inclinometro.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'  # Garante codificação UTF-8 no arquivo de log
)

def verificar_arquivo_excel():
    """Verifica se o arquivo Inclinômetro.xlsx existe no diretório especificado"""
    caminho_arquivo = os.path.join('media', 'planilhas_SLU', 'Inclinômetro.xlsx')
    logging.info(f"Verificando existência do arquivo: {caminho_arquivo}")
    
    if not os.path.exists(caminho_arquivo):
        logging.error(f"Arquivo não encontrado: {caminho_arquivo}")
        return False
        
    logging.info(f"Arquivo encontrado: {caminho_arquivo}")
    return True

def verificar_estrutura_excel(wb):
    """Verifica se o arquivo Excel tem as abas necessárias"""
    abas_necessarias = ['Coordenada NORTE', 'Coordenada ESTE', 'Cota']
    abas_existentes = wb.sheetnames
    
    logging.info(f"Abas encontradas no arquivo Excel: {abas_existentes}")
    
    for aba in abas_necessarias:
        if aba not in abas_existentes:
            logging.error(f"Aba não encontrada: {aba}")
            return False
    
    logging.info("Todas as abas necessárias foram encontradas")
    return True

def extrair_data_do_arquivo(nome_arquivo):
    """Extrai a data do nome do arquivo no formato especificado
    Procura por padrão DDMMAA entre a letra L no início e M no final
    Exemplo: L091023M para 09/10/2023"""
    logging.info(f"Tentando extrair data do arquivo: {nome_arquivo}")
    try:
        # Procura por padrão DDMMAA entre a letra L e M
        padrao = r'L(\d{6})M'
        match = re.search(padrao, nome_arquivo)
        if match:
            data_str = match.group(1)
            dia = int(data_str[:2])
            mes = int(data_str[2:4])
            ano = int(data_str[4:])
            
            # Validação básica das partes da data
            if dia < 1 or dia > 31 or mes < 1 or mes > 12:
                logging.error(f"Valores inválidos para dia ({dia}) ou mês ({mes}) no arquivo: {nome_arquivo}")
                return None
                
            # Assumindo que anos 00-50 são 2000-2050
            ano_completo = 2000 + ano if ano < 50 else 1900 + ano
            
            try:
                data = datetime(ano_completo, mes, dia)
                logging.info(f"Data extraída com sucesso: {data.strftime('%d/%m/%Y')}")
                logging.info(f"Partes da data - Dia: {dia}, Mês: {mes}, Ano: {ano_completo}")
                return data
            except ValueError as ve:
                logging.error(f"Data inválida no arquivo {nome_arquivo}: {str(ve)}")
                return None
        else:
            logging.error(f"Padrão de data (LDDMMAAM) não encontrado no arquivo: {nome_arquivo}")
            # Mostrar a parte do nome do arquivo que contém a data para debug
            logging.info(f"Conteúdo do arquivo onde deveria estar a data: {re.findall(r'L\d*M', nome_arquivo)}")
    except Exception as e:
        logging.error(f"Erro ao extrair data do arquivo {nome_arquivo}: {str(e)}\n{traceback.format_exc()}")
    return None

def calcular_dias_ate_referencia(data):
    """Calcula o número de dias entre a data fornecida e 13/04/2016"""
    data_referencia = datetime(2016, 4, 13)
    dias = (data - data_referencia).days
    logging.info(f"Dias calculados entre {data.strftime('%d/%m/%Y')} e 13/04/2016: {dias}")
    return dias

def formatar_numero_br(valor):
    """Formata número para o padrão brasileiro com 3 casas decimais"""
    try:
        if pd.isna(valor):
            logging.warning(f"Valor nulo encontrado")
            return None
        numero_formatado = locale.format_string("%.3f", float(valor), grouping=True)
        logging.debug(f"Valor formatado: {valor} -> {numero_formatado}")
        return numero_formatado
    except Exception as e:
        logging.error(f"Erro ao formatar número {valor}: {str(e)}")
        return None

def processar_dados():
    """Função principal de processamento dos dados"""
    logging.info("Iniciando processamento dos dados")
    
    if not verificar_arquivo_excel():
        return
    
    try:
        # Conectar ao banco de dados
        logging.info("Tentando conectar ao banco de dados...")
        if not os.path.exists('banco_dados.db'):
            logging.error("Arquivo do banco de dados não encontrado: banco_dados.db")
            return
            
        conn = sqlite3.connect('banco_dados.db')
        logging.info("Conexão com banco de dados estabelecida com sucesso")
        
        # Verificar se a tabela existe e sua estrutura
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='placas_completas_slu_bh'")
        if not cursor.fetchone():
            logging.error("Tabela 'placas_completas_slu_bh' não encontrada no banco de dados")
            conn.close()
            return
            
        # Verificar estrutura da tabela
        cursor.execute("PRAGMA table_info(placas_completas_slu_bh)")
        colunas = cursor.fetchall()
        logging.info("Estrutura da tabela:")
        for coluna in colunas:
            logging.info(f"Coluna: {coluna[1]}, Tipo: {coluna[2]}")
            
        # Verificar quantidade total de registros na tabela
        cursor.execute("SELECT COUNT(*) FROM placas_completas_slu_bh")
        total_registros_tabela = cursor.fetchone()[0]
        logging.info(f"Total de registros na tabela: {total_registros_tabela}")
            
        # Consultar registros de inclinômetro
        logging.info("Executando consulta SQL...")
        # Consulta para ver os tipos distintos presentes no banco
        cursor.execute("SELECT DISTINCT tipo FROM placas_completas_slu_bh")
        tipos = cursor.fetchall()
        logging.info(f"Tipos encontrados no banco: {tipos}")
        
        query = """
        SELECT nome_arquivo, info_gerada, coordenada_norte, coordenada_este, elevacao 
        FROM placas_completas_slu_bh 
        WHERE tipo = 'inclinômetro'
        """
        
        df = pd.read_sql_query(query, conn)
        total_registros = len(df)
        logging.info(f"Total de registros encontrados no banco: {total_registros}")
        
        if total_registros == 0:
            logging.warning("Nenhum registro de inclinômetro encontrado no banco")
            conn.close()
            return
            
        conn.close()
        logging.info("Conexão com banco de dados fechada")
        
        # Carregar arquivo Excel
        excel_path = os.path.join('media', 'planilhas_SLU', 'Inclinômetro.xlsx')
        logging.info(f"Carregando arquivo Excel: {excel_path}")
        wb = load_workbook(excel_path)
        
        if not verificar_estrutura_excel(wb):
            return
        
        # Dicionário para mapear info_gerada para colunas Excel
        mapa_colunas = {
            'I8': 'C', 'I7': 'D', 'I3': 'E', 'I2': 'F', 'I1': 'G',
            'I6': 'H', 'I4': 'I', 'I5': 'J'
        }
        logging.info(f"Mapeamento de colunas: {mapa_colunas}")
        
        # Contadores para relatório
        registros_processados = 0
        registros_duplicados = 0
        registros_com_erro = 0
        
        # Processar cada registro
        registros_por_data = {}
        for idx, row in df.iterrows():
            logging.info(f"\nProcessando registro {idx + 1} de {total_registros}")
            logging.info(f"Nome do arquivo: {row['nome_arquivo']}")
            
            try:
                # Extrair data do nome do arquivo
                data = extrair_data_do_arquivo(row['nome_arquivo'])
                if not data:
                    registros_com_erro += 1
                    continue
                
                # Verificar se já processamos esta data
                if data not in registros_por_data:
                    dias_corridos = calcular_dias_ate_referencia(data)
                    registros_por_data[data] = {
                        'norte': {},
                        'este': {},
                        'cota': {}
                    }
                    logging.info(f"Nova data encontrada: {data.strftime('%d/%m/%Y')}, dias corridos: {dias_corridos}")
                
                info_gerada = row['info_gerada']
                if info_gerada not in mapa_colunas:
                    logging.warning(f"info_gerada inválida: {info_gerada}")
                    registros_com_erro += 1
                    continue
                
                coluna = mapa_colunas[info_gerada]
                logging.info(f"info_gerada: {info_gerada}, coluna Excel: {coluna}")
                
                # Formatar valores
                norte = formatar_numero_br(row['coordenada_norte'])
                este = formatar_numero_br(row['coordenada_este'])
                cota = formatar_numero_br(row['elevacao'])
                
                logging.info(f"Valores formatados - Norte: {norte}, Este: {este}, Cota: {cota}")
                
                # Armazenar valores formatados
                registros_por_data[data]['norte'][coluna] = norte
                registros_por_data[data]['este'][coluna] = este
                registros_por_data[data]['cota'][coluna] = cota
                
                registros_processados += 1
                
            except Exception as e:
                logging.error(f"Erro ao processar registro: {str(e)}\n{traceback.format_exc()}")
                registros_com_erro += 1
                continue
        
        # Gravar dados no Excel
        logging.info("\nIniciando gravação no Excel")
        for data, valores in registros_por_data.items():
            data_str = data.strftime('%d/%m/%Y')
            dias = calcular_dias_ate_referencia(data)
            
            for tipo_planilha in ['Coordenada NORTE', 'Coordenada ESTE', 'Cota']:
                sheet = wb[tipo_planilha]
                logging.info(f"\nProcessando aba: {tipo_planilha}")
                
                # Encontrar primeira linha vazia
                ultima_linha = 1
                while sheet[f'A{ultima_linha}'].value is not None:
                    ultima_linha += 1
                
                logging.info(f"Primeira linha vazia encontrada: {ultima_linha}")
                
                # Gravar data e dias
                sheet[f'A{ultima_linha}'] = data_str
                sheet[f'B{ultima_linha}'] = dias
                logging.info(f"Data {data_str} e dias {dias} gravados na linha {ultima_linha}")
                
                # Gravar valores nas colunas apropriadas
                dados = valores['norte' if tipo_planilha == 'Coordenada NORTE' else 
                              'este' if tipo_planilha == 'Coordenada ESTE' else 'cota']
                
                for coluna, valor in dados.items():
                    sheet[f'{coluna}{ultima_linha}'] = valor
                    logging.info(f"Valor {valor} gravado na célula {coluna}{ultima_linha}")
        
        # Salvar arquivo
        logging.info(f"\nSalvando arquivo Excel: {excel_path}")
        wb.save(excel_path)
        logging.info("Arquivo Excel salvo com sucesso")
        
        # Gerar relatório final
        relatorio = f"""
        Processamento concluído:
        - Total de registros encontrados: {total_registros}
        - Registros processados com sucesso: {registros_processados}
        - Registros duplicados ignorados: {registros_duplicados}
        - Registros com erro: {registros_com_erro}
        """
        logging.info(relatorio)
        
    except Exception as e:
        logging.error(f"Erro durante o processamento: {str(e)}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    try:
        # Configurar locale
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
            logging.info("Locale configurado para pt_BR.UTF-8")
        except locale.Error:
            try:
                locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
                logging.info("Locale configurado para Portuguese_Brazil.1252")
            except locale.Error:
                logging.error("Não foi possível configurar o locale para português do Brasil")
                raise
        
        processar_dados()
    except Exception as e:
        logging.error(f"Erro fatal durante execução: {str(e)}\n{traceback.format_exc()}")