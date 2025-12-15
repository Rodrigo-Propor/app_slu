import sqlite3
import os
from datetime import datetime, date
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import re

################################################################################
#                          FUNÇÕES DE APOIO                                    #
################################################################################

def obter_conexao_banco(caminho_banco):
    """
    Abre (ou cria) a conexão com o banco de dados SQLite no caminho especificado.
    Retorna o objeto de conexão, ou None em caso de erro.
    """
    try:
        conn = sqlite3.connect(caminho_banco)
        return conn
    except Exception as e:
        st.error(f"Não foi possível conectar ao banco: {e}", icon="🚫")
        st.exception(e)
        return None

def buscar_ultimas_datas_por_arquivo_planilha(conn):
    """
    Lê na tabela 'ultima_linha_arquivos' as informações de última data
    de cada arquivo e planilha. Retorna um dicionário no formato:
        {
          (nome_arquivo, nome_planilha): (linha_informacao, data_datetime)
        }
    Se não houver registro, retorna dicionário vazio.
    """
    d = {}
    try:
        # Verificar se a tabela existe
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ultima_linha_arquivos'")
        if not cur.fetchone():
            st.warning("A tabela 'ultima_linha_arquivos' não existe. Usando valores padrão.", icon="⚠️")
            return d
            
        cur.execute("""
            SELECT nome_arquivo, nome_planilha, linha_informacao, data_ultimo_registro
            FROM ultima_linha_arquivos
        """)
        rows = cur.fetchall()
        for (arq, plan, linha, data_str) in rows:
            try:
                data_dt = datetime.strptime(data_str, "%d/%m/%Y").date()
            except:
                try:
                    data_dt = datetime.strptime(data_str, "%Y-%m-%d").date()
                except:
                    data_dt = date(1900, 1, 1)
            d[(arq, plan)] = (linha, data_dt)
    except Exception as e:
        st.error("Erro ao buscar últimos registros no banco de dados:", icon="🚫")
        st.exception(e)
    return d

def converter_str_para_date(data_str):
    """
    Converte a coluna 'data' do banco (em formato yyyy-mm-dd ou dd/mm/aaaa) para um objeto date.
    Retorna None se não conseguir converter.
    """
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(data_str, fmt).date()
        except ValueError:
            pass
    return None

def calcular_dias_desde_referencia(data_linha, data_referencia):
    """
    Calcula quantos dias se passaram entre data_linha e data_referencia.
    Retorna inteiro (pode ser negativo se data_linha < data_referencia).
    """
    return (data_linha - data_referencia).days

def ajustar_formatacao_celula(celula, eh_data=False):
    """
    Ajusta a formatação da célula segundo padrão brasileiro (pt-BR).
    - Se eh_data=True: tenta formatar a célula como data (dd-mmm-yy).
    - Caso contrário, formata com 3 casas decimais e usa vírgula como decimal (se o Excel permitir).
    """
    if eh_data:
        celula.number_format = 'dd-mmm-yy'  # Ex.: 30-ago-24 (dependendo do locale do Office)
    else:
        celula.number_format = '#,##0.000'
    return

def definir_arquivo_e_planilha(tipo, placa):
    """
    Dada a coluna 'tipo' e 'placa', retorna (arquivo_destino, planilha_destino) segundo as regras:
      - CELULA EMERGENCIAL
      - INCLINOMETRO
      - CELULA DE PESQUISA
      - PAMPULHA
    Caso não se encaixe em nenhuma regra, retorna (None, None).
    """
    placa_strip = placa.strip().upper()
    tipo_up = tipo.strip().upper()
    
    # 1) CELULA EMERGENCIAL
    if tipo_up == "CELULA EMERGENCIAL":
        # Verifica se placa começa com "PR "
        if placa_strip.startswith("PR "):
            arquivo = sanitize_filename("Placa de Recalque Emergencial.xlsx")
            planilha = placa.strip()  # ex.: "PR 3.4"
            return (arquivo, planilha)
        else:
            return (None, None)
    
    
    # 3) CELULA DE PESQUISA
    if tipo_up == "CELULA DE PESQUISA":
        # placa começa com "PR"
        if placa_strip.startswith("PR"):
            arquivo = sanitize_filename("Recalques Celula Pesquisa.xlsx")
            planilha = placa_strip  # "PR03"
            return (arquivo, planilha)
        else:
            return (None, None)
    
    # 4) PAMPULHA
    if tipo_up == "PAMPULHA":
        # placa que começa com "PR 1." ou é "D1" ou "D2" => AC 01
        if placa_strip.startswith("PR 1.") or placa_strip in ["D1", "D2"]:
            return (sanitize_filename("Placa de Recalque AC 01.xlsx"), placa.strip())
        # placa que começa com "PR 3." => AC 03
        if placa_strip.startswith("PR 3."):
            return (sanitize_filename("Placa de Recalque AC 03.xlsx"), placa.strip())
        # placa que começa com "PR 4." => AC 04
        if placa_strip.startswith("PR 4."):
            return (sanitize_filename("Placa de Recalque AC 04.xlsx"), placa.strip())
        # placa que começa com "PR 5." => AC 05
        if placa_strip.startswith("PR 5."):
            return (sanitize_filename("Placa de Recalque AC 05.xlsx"), placa.strip())
        # placa que começa com "PR A." => AMPLIAÇÃO
        if placa_strip.startswith("PR A."):
            return (sanitize_filename("Placa de Recalque AMPLIAÇÃO.xlsx"), placa.strip())
        return (None, None)
    
    return (None, None)

def encontrar_ultima_linha_valida_banco(conn, nome_arquivo, nome_planilha, ws):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT ultima_linha_valida
        FROM verificacao_planilhas
        WHERE nome_arquivo = ? AND nome_planilha = ?
        ORDER BY data_processamento DESC
        LIMIT 1
    """, (nome_arquivo, nome_planilha))
    row = cursor.fetchone()
    if row:
        return row[0]
    ultima_linha = ws.max_row
    while ultima_linha > 0 and ws.cell(ultima_linha, 1).value is None:
        ultima_linha -= 1
    return ultima_linha

def inserir_registro_padrao(ws, conn, arquivo, planilha, data_reg, coord_este, coord_norte, elevacao):
    """
    Insere um novo registro no Excel na próxima linha disponível.
    """
    # Determinar a última linha preenchida na planilha
    ultima_linha_valida = ws.max_row
    while ultima_linha_valida > 0 and ws.cell(ultima_linha_valida, 1).value is None:
        ultima_linha_valida -= 1

    # Próxima linha disponível
    new_row = ultima_linha_valida + 1

    # Inserir os dados na nova linha
    ws.cell(new_row, 1).value = data_reg
    ajustar_formatacao_celula(ws.cell(new_row, 1), eh_data=True)

    ws.cell(new_row, 2).value = elevacao
    ajustar_formatacao_celula(ws.cell(new_row, 2), eh_data=False)

    ws.cell(new_row, 3).value = coord_este
    ajustar_formatacao_celula(ws.cell(new_row, 3), eh_data=False)

    ws.cell(new_row, 4).value = coord_norte
    ajustar_formatacao_celula(ws.cell(new_row, 4), eh_data=False)

def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

def criar_tabela_ultima_linha_if_not_exists(conn):
    """
    Cria a tabela 'ultima_linha_arquivos' no banco de dados se ela ainda não existir.
    """
    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ultima_linha_arquivos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome_arquivo TEXT,
                nome_planilha TEXT,
                linha_informacao INTEGER,
                data_ultimo_registro TEXT,
                data_processamento TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        conn.commit()
        return True
    except Exception as e:
        st.error(f"Erro ao criar tabela 'ultima_linha_arquivos': {e}", icon="🚫")
        st.exception(e)
        return False

def atualizar_ultima_linha_arquivos(conn, arquivo, planilha, linha, data_ultimo):
    """
    Atualiza (ou insere) na tabela ultima_linha_arquivos o registro de última
    linha processada e data do último registro inserido.
    """
    try:
        data_str = data_ultimo.strftime("%Y-%m-%d")
        cursor = conn.cursor()
        
        # Verificar se já existe um registro para este arquivo/planilha
        cursor.execute("""
            SELECT id FROM ultima_linha_arquivos
            WHERE nome_arquivo = ? AND nome_planilha = ?
        """, (arquivo, planilha))
        
        row = cursor.fetchone()
        if row:
            # Atualiza o registro existente
            cursor.execute("""
                UPDATE ultima_linha_arquivos
                SET linha_informacao = ?, data_ultimo_registro = ?, data_processamento = CURRENT_TIMESTAMP
                WHERE nome_arquivo = ? AND nome_planilha = ?
            """, (linha, data_str, arquivo, planilha))
        else:
            # Insere um novo registro
            cursor.execute("""
                INSERT INTO ultima_linha_arquivos 
                (nome_arquivo, nome_planilha, linha_informacao, data_ultimo_registro)
                VALUES (?, ?, ?, ?)
            """, (arquivo, planilha, linha, data_str))
        
        conn.commit()
        return True
    except Exception as e:
        st.error(f"Erro ao atualizar tabela de controle: {e}", icon="🚫")
        st.exception(e)
        return False

################################################################################
#                             CÓDIGO PRINCIPAL                                 #
################################################################################

def main():
    st.title("Exportação de Dados para Excel")
    
    # Ajuste conforme o seu cenário
    CAMINHO_BANCO = "banco_dados.db"
    DIRETORIO_EXCEL = "media/planilhas_SLU"
    
    # Adicionar botão para iniciar o processamento
    if st.button("Iniciar Processamento"):
        # 1) Conectar ao banco
        conn = obter_conexao_banco(CAMINHO_BANCO)
        if not conn:
            st.error("Não foi possível prosseguir sem conexão ao banco.", icon="🚫")
            return
        
        st.info("Conexão ao banco estabelecida com sucesso!", icon="ℹ️")
        
        # 1.1) Verificar e criar tabela de controle se necessário
        if not criar_tabela_ultima_linha_if_not_exists(conn):
            st.error("Não foi possível preparar as tabelas necessárias. Abortando processo.", icon="🚫")
            conn.close()
            return
            
        # 2) Obter as últimas datas registradas (controladas em 'ultima_linha_arquivos')
        ultimas_datas = buscar_ultimas_datas_por_arquivo_planilha(conn)
        st.info("Leitura das últimas datas registradas concluída.", icon="ℹ️")
        
        # 3) Buscar registros na tabela 'placas_completas_slu_bh'
        st.status("Buscando registros no banco...", state="running", expanded=False)
        try:
            cur = conn.cursor()
            # IMPORTANTE: Aqui corrigimos o nome da tabela!
            cur.execute("""
                SELECT id, data, tipo, descricao, coordenada_este, coordenada_norte, elevacao, placa 
                FROM placas_completas_slu_bh
            """)
            todos_registros = cur.fetchall()
        except Exception as e:
            st.error("Erro ao consultar a tabela 'placas_completas_slu_bh' no banco de dados!", icon="🚫")
            st.exception(e)
            return
        
        st.success(f"Foram lidos {len(todos_registros)} registros do banco.", icon="✅")
        
        # 4) Filtrar e classificar registros
        registros_validos = []
        for row in todos_registros:
            row_id, data_str, tipo, desc, este, norte, elev, placa = row
            data_dt = converter_str_para_date(data_str)
            if not data_dt:
                st.warning(f"O registro ID={row_id} tem data inválida: '{data_str}'. Ignorando...", icon="⚠️")
                continue
            
            # Descobrir destino
            arquivo_dest, planilha_dest = definir_arquivo_e_planilha(tipo, placa)
            if not arquivo_dest or not planilha_dest:
                st.warning(f"Registro ID={row_id} não se encaixa em nenhuma regra de destino. Ignorando...", icon="⚠️")
                continue
            
            # Verificar se data_dt > última data do (arquivo_dest, planilha_dest)
            ultima_data_conhecida, linha_informacao = ultimas_datas.get((arquivo_dest, planilha_dest), (0, date(1900,1,1)))
            if not isinstance(ultima_data_conhecida, date):
                ultima_data_conhecida = date(1900, 1, 1)
            if data_dt <= ultima_data_conhecida:
                # Não grava, pois não é posterior
                continue
            
            registros_validos.append({
                "id": row_id,
                "data_dt": data_dt,
                "tipo": tipo,
                "placa": placa,
                "coord_este": este,
                "coord_norte": norte,
                "elevacao": elev,
                "arquivo": arquivo_dest,
                "planilha": planilha_dest,
                "linha_informacao": linha_informacao
            })
        
        st.info(f"Temos {len(registros_validos)} registros que precisam ser gravados.", icon="ℹ️")
        if not registros_validos:
            st.warning("Nenhum registro a gravar. Encerrando...", icon="⚠️")
            conn.close()
            return
        
        # 5) Agrupar por (arquivo, planilha)
        from collections import defaultdict
        grupos = defaultdict(list)
        for reg in registros_validos:
            chave = (reg["arquivo"], reg["planilha"])
            grupos[chave].append(reg)
        
        # 6) Gravar em cada arquivo/planilha
        total_grupos = len(grupos)
        progresso_geral = st.progress(0, text="Iniciando gravação nos arquivos...")
        cont_grupos = 0
        
        for (arquivo_excel, planilha_excel), regs_grupo in grupos.items():
            cont_grupos += 1
            st.info(f"Iniciando processamento do arquivo '{arquivo_excel}' / planilha '{planilha_excel}'", icon="ℹ️")
            
            # O arquivo já deve estar sanitizado na função definir_arquivo_e_planilha, mas vamos garantir
            arquivo_excel = sanitize_filename(arquivo_excel)
            caminho_arq = os.path.join(DIRETORIO_EXCEL, arquivo_excel)
            
            try:
                if not os.path.exists(caminho_arq):
                    st.error(f"Arquivo '{arquivo_excel}' não foi encontrado. Não será possível gravar esses registros.", icon="🚫")
                    continue
                
                wb = load_workbook(caminho_arq)
                
                # Remover a parte do código que faz a transferência dos dados do Inclinômetro
                # if planilha_excel == "INCLINÔMETRO" and arquivo_excel == "Inclinômetro.xlsx":
                #     st.warning(f"Registros para Inclinômetro não serão processados. Ignorando...", icon="⚠️")
                #     continue
                
                # Planilha comum
                if planilha_excel not in wb.sheetnames:
                    st.error(f"Planilha '{planilha_excel}' não encontrada em '{arquivo_excel}'. Não gravado.", icon="🚫")
                    continue
                
                ws = wb[planilha_excel]
                regs_ordenados = sorted(regs_grupo, key=lambda x: x["data_dt"])
                
                with st.spinner(f"Gravando {len(regs_ordenados)} pontos em '{arquivo_excel}' / '{planilha_excel}'..."):
                    for reg in regs_ordenados:
                        inserir_registro_padrao(
                            ws,
                            conn,
                            arquivo_excel,
                            planilha_excel,
                            reg["data_dt"],
                            reg["coord_este"],
                            reg["coord_norte"],
                            reg["elevacao"]
                        )
                wb.save(caminho_arq)
                
                # Atualizar a tabela de controle com a última data processada
                ultima_linha = encontrar_ultima_linha_valida_banco(conn, arquivo_excel, planilha_excel, ws)
                data_ultimo_reg = regs_ordenados[-1]["data_dt"] if regs_ordenados else date.today()
                atualizar_ultima_linha_arquivos(conn, arquivo_excel, planilha_excel, ultima_linha, data_ultimo_reg)
                
                st.success(f"Registros gravados em '{arquivo_excel}' / '{planilha_excel}'.", icon="✅")
            except Exception as e:
                st.error(f"Erro ao processar arquivo '{arquivo_excel}':", icon="🚫")
                st.exception(e)
                continue
            
            progresso_geral.progress(int((cont_grupos / total_grupos) * 100),
                                     text=f"Progresso geral: {cont_grupos}/{total_grupos}")
        
        st.success("Processo de gravação concluído com êxito!", icon="✅")
        conn.close()
        st.info("Conexão ao banco encerrada.", icon="ℹ️")

if __name__ == "__main__":
    main()
