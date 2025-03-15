import openpyxl
from openpyxl.styles import numbers
import os
from datetime import datetime
import streamlit as st
import pandas as pd

def parse_date(value):
    """Converte strings no formato 'dd-mmm-aa' para datetime"""
    months = {
        "jan": "Jan", "fev": "Feb", "mar": "Mar", "abr": "Apr", "mai": "May", "jun": "Jun",
        "jul": "Jul", "ago": "Aug", "set": "Sep", "out": "Oct", "nov": "Nov", "dez": "Dec"
    }
    try:
        if isinstance(value, str):
            for pt, en in months.items():
                value = value.replace(f"-{pt}-", f"-{en}-")
            return datetime.strptime(value, "%d-%b-%y")
    except ValueError:
        return value
    return value

def parse_number(value, column):
    """Converte strings para números, ajustando escala quando necessário"""
    try:
        if isinstance(value, str):
            value = value.replace('.', '').replace(',', '.')
            number = float(value)

            if column == 'B' and number > 999:
                return number / 1000
            elif column == 'C' and number > 999999:
                return number / 1000
            elif column == 'D' and number > 9999999:
                return number / 1000
            return number
    except ValueError:
        return value
    return value

def find_last_row(sheet, column='A'):
    """Encontra a última linha com dados"""
    max_row = sheet.max_row
    for row in range(max_row, 0, -1):
        cell_value = sheet[f'{column}{row}'].value
        if cell_value is not None:
            next_5_empty = all(
                sheet[f'{column}{row + i}'].value is None
                for i in range(1, 6)
            )
            if next_5_empty:
                return row
    return 1

def get_last_data_rows(sheet, count=20):
    """Retorna uma lista com os índices das últimas 'count' linhas que possuem dados
    nas colunas B, C ou D."""
    data_rows = []
    for row in range(sheet.max_row, 0, -1):
        # Verifica se alguma das colunas B, C ou D possui valor
        if any(sheet.cell(row=row, column=col).value is not None for col in [2, 3, 4]):
            data_rows.append(row)
        if len(data_rows) == count:
            break
    return sorted(data_rows)

def format_cells(workbook, sheet):
    """Formata as células das últimas 20 linhas com dados de coordenadas e cotas e retorna as linhas formatadas"""
    rows_to_format = get_last_data_rows(sheet, count=20)
    
    for row in rows_to_format:
        if row <= 0:
            continue
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            col_letter = chr(col + 64)
            if isinstance(cell.value, str):
                if col_letter in ['B', 'C', 'D']:
                    # Força a conversão para número
                    cell.value = parse_number(cell.value, col_letter)
                else:
                    cell.value = parse_date(cell.value) or parse_number(cell.value, col_letter) or cell.value
            
            # Formatação específica para cada tipo de dado
            if isinstance(cell.value, datetime):
                # Formato personalizado para data: dd/mmm/yy
                cell.number_format = 'dd/mmm/yy'
            elif isinstance(cell.value, (int, float)):
                # Mantém o formato para números com 3 casas decimais
                cell.number_format = '0.000'
    
    return rows_to_format

def main():
    st.title("Formatação de Planilhas Excel")
    
    # Define o caminho da pasta de planilhas
    pasta_planilhas = os.path.join('media', 'planilhas_SLU')
    
    # Verifica se a pasta existe
    if not os.path.exists(pasta_planilhas):
        st.error(f"A pasta {pasta_planilhas} não foi encontrada!")
        return
    
    # Lista arquivos Excel no diretório específico
    excel_files = [arquivo for arquivo in os.listdir(pasta_planilhas) if arquivo.endswith('.xlsx')]
    
    if not excel_files:
        st.warning(f"Nenhum arquivo Excel (.xlsx) encontrado em {pasta_planilhas}!")
        return
    
    st.write("Arquivos Excel encontrados:")
    for file in excel_files:
        st.text(f"📊 {file}")
    
    # Lista para armazenar as informações das linhas formatadas
    format_info = []
    
    if st.button("Iniciar Formatação de Todos os Arquivos"):
        total_files = len(excel_files)
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for file_index, file_name in enumerate(excel_files):
            try:
                # Usa o caminho completo do arquivo
                file_path = os.path.join(pasta_planilhas, file_name)
                workbook = openpyxl.load_workbook(file_path)
                total_sheets = len(workbook.sheetnames)
                
                for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                    status_text.text(f"Processando {file_name} - Aba: {sheet_name}")
                    sheet = workbook[sheet_name]
                    formatted_rows = format_cells(workbook, sheet)
                    
                    # Armazena a informação de cada linha formatada
                    for row in formatted_rows:
                        format_info.append({
                            "Arquivo": file_name,
                            "Aba": sheet_name,
                            "Linha Formatada": row
                        })
                    
                    # Calcula o progresso total (arquivos e abas)
                    progress = (file_index + (sheet_index + 1) / total_sheets) / total_files
                    progress_bar.progress(progress)
                
                workbook.save(file_path)
                
            except Exception as e:
                st.error(f"Erro ao processar {file_name}: {str(e)}")
        
        progress_bar.progress(1.0)
        status_text.text("Processamento concluído!")
        st.success("Todos os arquivos foram processados com sucesso!")
        
        # Exibe DataFrame com as informações das linhas formatadas
        if format_info:
            df_info = pd.DataFrame(format_info)
            st.write("Linhas formatadas:")
            st.dataframe(df_info)

if __name__ == "__main__":
    main()
