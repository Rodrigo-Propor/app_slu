import streamlit as st
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime, date

###############################################################################
#                            FUNÇÕES DE APOIO                                 #
###############################################################################

def obter_conexao_banco(caminho_banco):
    """
    Abre e retorna uma conexão com o banco SQLite. 
    Caso não consiga, retorna None.
    """
    try:
        conn = sqlite3.connect(caminho_banco)
        return conn
    except Exception as e:
        st.error(f"Erro ao conectar ao banco: {e}")
        return None

def converter_data_para_date(data_str):
    """
    Tenta converter a string de data para um objeto date, 
    usando vários formatos. Retorna None se não conseguir.
    """
    formatos = [
        "%Y-%m-%d",  # 2025-02-14
        "%d/%m/%Y",  # 14/02/2025
        "%d-%m-%Y",  # 14-02-2025
        "%d-%b-%y",  # 14-fev-25
        "%d-%b-%Y",  # 14-fev-2025
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(data_str, fmt).date()
        except ValueError:
            pass
    return None

def calcular_dias(data_linha, data_ref=date(2016,4,13)):
    """
    Retorna quantos dias se passaram entre data_linha e 13/04/2016.
    """
    return (data_linha - data_ref).days

def ajustar_formatacao_celula(celula, eh_data=False):
    """
    Define a formatação 'pt-BR' (tentativa):
     - Se eh_data=True, 'dd-mmm-yy' (Ex.: 14-fev-25).
     - Caso contrário, '#,##0.000' para usar vírgula decimal e 3 casas.
    """
    if eh_data:
        celula.number_format = 'dd-mmm-yy'
    else:
        celula.number_format = '#,##0.000'

###############################################################################
#                         CÓDIGO PRINCIPAL                                    #
###############################################################################

def main():
    st.title("Exportação para Inclinômetro.xlsx – 3 planilhas separadas")
    
    if st.button("Iniciar Processamento"):
        # 1. Caminhos
        CAMINHO_BANCO = r"D:\OneDrive\Documentos\APP\app_slu\banco_dados.db"
        DIRETORIO_EXCEL = r"D:\OneDrive\Documentos\APP\app_slu\media\planilhas_SLU"
        NOME_ARQUIVO = "Inclinômetro.xlsx"
        
        # 2. Conexão ao banco
        conn = obter_conexao_banco(CAMINHO_BANCO)
        if not conn:
            return
        
        # 3. Ler registros de INCLINÔMETRO
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT id, data, tipo, descricao, coordenada_este, coordenada_norte, elevacao, placa
                FROM placas_completas_slu_bh
                WHERE tipo = 'INCLINOMETRO'
            """)
            rows = cur.fetchall()
        except Exception as e:
            st.error("Erro ao consultar registros de INCLINÔMETRO:", icon="🚫")
            st.exception(e)
            conn.close()
            return
        
        st.success(f"Foram carregados {len(rows)} registros de INCLINÔMETRO.", icon="✅")
        if not rows:
            conn.close()
            return
        
        # 4. Agrupar por data, depois por placa
        #    Precisamos armazenar TANTO a coordenada_norte, ESTE quanto a ELEVACAO 
        #    para cada data e placa.
        #    Estrutura: 
        #    dados_por_data[ data ] [ placa ] = (coordenada_este, coordenada_norte, elevacao)
        dados_por_data = {}
        
        for row in rows:
            (row_id, data_str, tipo, descricao, este, norte, elev, placa) = row
            data_dt = converter_data_para_date(data_str)
            if not data_dt:
                continue
            
            placa_up = placa.strip().upper()
            
            if data_dt not in dados_por_data:
                dados_por_data[data_dt] = {}
            # Armazena os 3 valores
            dados_por_data[data_dt][placa_up] = (este, norte, elev)
        
        # 5. Ordenar as datas
        lista_datas_ordenadas = sorted(dados_por_data.keys())
        
        # 6. Abrir o arquivo Inclinômetro.xlsx
        caminho_xlsx = os.path.join(DIRETORIO_EXCEL, NOME_ARQUIVO)
        if not os.path.exists(caminho_xlsx):
            st.error(f"Arquivo '{NOME_ARQUIVO}' não encontrado em '{DIRETORIO_EXCEL}'.")
            conn.close()
            return
        
        wb = load_workbook(caminho_xlsx)
        
        # Verifica se as 3 planilhas existem
        abas_necessarias = ["Coordenada NORTE", "Coordenada ESTE", "Cota"]
        for aba in abas_necessarias:
            if aba not in wb.sheetnames:
                st.error(f"Planilha '{aba}' não existe em '{NOME_ARQUIVO}'.")
                conn.close()
                return
        
        ws_norte = wb["Coordenada NORTE"]
        ws_este  = wb["Coordenada ESTE"]
        ws_cota  = wb["Cota"]
        
        st.info(f"Inserindo dados em {len(lista_datas_ordenadas)} datas distintas.", icon="ℹ️")
        
        # Mapeamento placa -> coluna
        map_placa_col = {
            "I8": 3,
            "I7": 4,
            "I3": 5,
            "I2": 6,
            "I1": 7,
            "I6": 8,
            "I4": 9,
            "I5": 10
        }
        
        # 7. Inserir em cada planilha 1 linha por data
        #    - ws_norte => coordenada_norte
        #    - ws_este  => coordenada_este
        #    - ws_cota  => elevacao
        
        progresso = st.progress(0, text="Iniciando gravação...")
        total_datas = len(lista_datas_ordenadas)
        contador = 0
        
        for data_dt in lista_datas_ordenadas:
            contador += 1
            # Calcula dias
            dias = calcular_dias(data_dt)
            
            # Cria nova linha na planilha NORTE
            row_norte = ws_norte.max_row + 1
            ws_norte.cell(row_norte, 1).value = data_dt
            ajustar_formatacao_celula(ws_norte.cell(row_norte, 1), eh_data=True)
            
            ws_norte.cell(row_norte, 2).value = dias
            ajustar_formatacao_celula(ws_norte.cell(row_norte, 2), eh_data=False)
            
            # Cria nova linha na planilha ESTE
            row_este = ws_este.max_row + 1
            ws_este.cell(row_este, 1).value = data_dt
            ajustar_formatacao_celula(ws_este.cell(row_este, 1), eh_data=True)
            
            ws_este.cell(row_este, 2).value = dias
            ajustar_formatacao_celula(ws_este.cell(row_este, 2), eh_data=False)
            
            # Cria nova linha na planilha COTA
            row_cota = ws_cota.max_row + 1
            ws_cota.cell(row_cota, 1).value = data_dt
            ajustar_formatacao_celula(ws_cota.cell(row_cota, 1), eh_data=True)
            
            ws_cota.cell(row_cota, 2).value = dias
            ajustar_formatacao_celula(ws_cota.cell(row_cota, 2), eh_data=False)
            
            # Para cada placa daquela data, colocar o valor certo nas 3 planilhas
            dict_placas = dados_por_data[data_dt]
            
            for placa_up, tupla_vals in dict_placas.items():
                # tupla_vals = (este, norte, elev)
                este_val, norte_val, elev_val = tupla_vals
                
                if placa_up in map_placa_col:
                    col = map_placa_col[placa_up]
                    
                    # NORTE => grava 'norte_val'
                    ws_norte.cell(row_norte, col).value = norte_val
                    ajustar_formatacao_celula(ws_norte.cell(row_norte, col), eh_data=False)
                    
                    # ESTE => grava 'este_val'
                    ws_este.cell(row_este, col).value = este_val
                    ajustar_formatacao_celula(ws_este.cell(row_este, col), eh_data=False)
                    
                    # COTA => grava 'elev_val'
                    ws_cota.cell(row_cota, col).value = elev_val
                    ajustar_formatacao_celula(ws_cota.cell(row_cota, col), eh_data=False)
            
            progresso.progress(int((contador / total_datas)*100),
                               text=f"Inserindo data {contador}/{total_datas}")
        
        # 8. Salvar
        wb.save(caminho_xlsx)
        st.success(f"{total_datas} registros de data gravados nas 3 planilhas!", icon="✅")
        
        # Fecha o banco
        conn.close()
        st.info("Conexão ao banco encerrada.", icon="ℹ️")

if __name__ == "__main__":
    main()
