# verificação_registros_excel.py

import streamlit as st
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import pandas as pd

# Caso precise fazer parse manual de datas em pt-BR (ex: "09/set/24"):
# import locale
# locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # ou "pt_BR" dependendo do sistema

# DIRETÓRIOS e ARQUIVOS
DIRETORIO_ARQUIVOS = "media/planilhas_SLU"
BANCO_DADOS = "banco_dados.db"

# Lista de arquivos e planilhas associadas:
# Aqui você pode mapear o 'id' à tupla (nome_arquivo, nome_planilha)
# ou simplesmente criar uma lista de dicionários, etc.
PLANILHAS_MAPEAMENTO = [
    # (id, nome_arquivo, nome_planilha)
    (1,   "Inclinômetro.xlsx",                "Coordenada NORTE"),
    (2,   "Inclinômetro.xlsx",                "Coordenada ESTE"),
    (3,   "Inclinômetro.xlsx",                "Cota"),
    (4,   "Placa de Recalque AC 01.xlsx",     "D1"),
    (5,   "Placa de Recalque AC 01.xlsx",     "D2"),
    (6,   "Placa de Recalque AC 01.xlsx",     "PR 1.1"),
    (7,   "Placa de Recalque AC 01.xlsx",     "PR 1.2"),
    (8,   "Placa de Recalque AC 01.xlsx",     "PR 1.3"),
    (9,   "Placa de Recalque AC 01.xlsx",     "PR 1.4"),
    (10,  "Placa de Recalque AC 01.xlsx",     "PR 1.5"),
    (11,  "Placa de Recalque AC 01.xlsx",     "PR 1.21"),
    (12,  "Placa de Recalque AC 01.xlsx",     "PR 1.22"),
    (13,  "Placa de Recalque AC 01.xlsx",     "PR 1.23"),
    (14,  "Placa de Recalque AC 01.xlsx",     "PR 1.24"),
    (15,  "Placa de Recalque AC 01.xlsx",     "PR 1.25"),
    (16,  "Placa de Recalque AC 01.xlsx",     "PR 1.31"),
    (17,  "Placa de Recalque AC 01.xlsx",     "PR 1.32"),
    (18,  "Placa de Recalque AC 01.xlsx",     "PR 1.33"),
    (19,  "Placa de Recalque AC 01.xlsx",     "PR 1.34"),
    (20,  "Placa de Recalque AC 01.xlsx",     "PR 1.35"),
    (21,  "Placa de Recalque AC 01.xlsx",     "PR 1.41"),
    (22,  "Placa de Recalque AC 01.xlsx",     "PR 1.42"),
    (23,  "Placa de Recalque AC 01.xlsx",     "PR 1.43"),
    (24,  "Placa de Recalque AC 01.xlsx",     "PR 1.44"),
    (25,  "Placa de Recalque AC 01.xlsx",     "PR 1.45"),
    (26,  "Placa de Recalque AC 03.xlsx",     "PR 3.21"),
    (27,  "Placa de Recalque AC 03.xlsx",     "PR 3.22"),
    (28,  "Placa de Recalque AC 03.xlsx",     "PR 3.23"),
    (29,  "Placa de Recalque AC 03.xlsx",     "PR 3.24"),
    (30,  "Placa de Recalque AC 03.xlsx",     "PR 3.25"),
    (31,  "Placa de Recalque AC 03.xlsx",     "PR 3.26"),
    (32,  "Placa de Recalque AC 03.xlsx",     "PR 3.27"),
    (33,  "Placa de Recalque AC 03.xlsx",     "PR 3.28"),
    (34,  "Placa de Recalque AC 03.xlsx",     "PR 3.29"),
    (35,  "Placa de Recalque AC 03.xlsx",     "PR 3.30"),
    (36,  "Placa de Recalque AC 03.xlsx",     "PR 3.31"),
    (37,  "Placa de Recalque AC 04.xlsx",     "PR 4.1"),
    (38,  "Placa de Recalque AC 04.xlsx",     "PR 4.2"),
    (39,  "Placa de Recalque AC 04.xlsx",     "PR 4.3"),
    (40,  "Placa de Recalque AC 04.xlsx",     "PR 4.4"),
    (41,  "Placa de Recalque AC 04.xlsx",     "PR 4.5"),
    (42,  "Placa de Recalque AC 04.xlsx",     "PR 4.6"),
    (43,  "Placa de Recalque AC 04.xlsx",     "PR 4.7"),
    (44,  "Placa de Recalque AC 04.xlsx",     "PR 4.8"),
    (45,  "Placa de Recalque AC 04.xlsx",     "PR 4.9"),
    (46,  "Placa de Recalque AC 04.xlsx",     "PR 4.10"),
    (47,  "Placa de Recalque AC 04.xlsx",     "PR 4.11"),
    (48,  "Placa de Recalque AC 04.xlsx",     "PR 4.21"),
    (49,  "Placa de Recalque AC 04.xlsx",     "PR 4.22"),
    (50,  "Placa de Recalque AC 04.xlsx",     "PR 4.23"),
    (51,  "Placa de Recalque AC 04.xlsx",     "PR 4.24"),
    (52,  "Placa de Recalque AC 04.xlsx",     "PR 4.25"),
    (53,  "Placa de Recalque AC 04.xlsx",     "PR 4.26"),
    (54,  "Placa de Recalque AC 04.xlsx",     "PR 4.27"),
    (55,  "Placa de Recalque AC 04.xlsx",     "PR 4.28"),
    (56,  "Placa de Recalque AC 04.xlsx",     "PR 4.29"),
    (57,  "Placa de Recalque AC 04.xlsx",     "PR 4.30"),
    (58,  "Placa de Recalque AC 04.xlsx",     "PR 4.31"),
    (59,  "Placa de Recalque AC 04.xlsx",     "PR 4.41"),
    (60,  "Placa de Recalque AC 04.xlsx",     "PR 4.42"),
    (61,  "Placa de Recalque AC 04.xlsx",     "PR 4.43"),
    (62,  "Placa de Recalque AC 04.xlsx",     "PR 4.44"),
    (63,  "Placa de Recalque AC 04.xlsx",     "PR 4.45"),
    (64,  "Placa de Recalque AC 04.xlsx",     "PR 4.46"),
    (65,  "Placa de Recalque AC 04.xlsx",     "PR 4.47"),
    (66,  "Placa de Recalque AC 04.xlsx",     "PR 4.48"),
    (67,  "Placa de Recalque AC 04.xlsx",     "PR 4.49"),
    (68,  "Placa de Recalque AC 04.xlsx",     "PR 4.50"),
    (69,  "Placa de Recalque AC 04.xlsx",     "PR 4.51"),
    (70,  "Placa de Recalque AC 04.xlsx",     "PR 4.61"),
    (71,  "Placa de Recalque AC 04.xlsx",     "PR 4.62"),
    (72,  "Placa de Recalque AC 04.xlsx",     "PR 4.63"),
    (73,  "Placa de Recalque AC 04.xlsx",     "PR 4.64"),
    (74,  "Placa de Recalque AC 04.xlsx",     "PR 4.65"),
    (75,  "Placa de Recalque AC 04.xlsx",     "PR 4.66"),
    (76,  "Placa de Recalque AC 04.xlsx",     "PR 4.67"),
    (77,  "Placa de Recalque AC 04.xlsx",     "PR 4.68"),
    (78,  "Placa de Recalque AC 04.xlsx",     "PR 4.69"),
    (79,  "Placa de Recalque AC 04.xlsx",     "PR 4.70"),
    (80,  "Placa de Recalque AC 04.xlsx",     "PR 4.71"),
    (81,  "Placa de Recalque AC 05.xlsx",     "PR 5.1"),
    (82,  "Placa de Recalque AC 05.xlsx",     "PR 5.2"),
    (83,  "Placa de Recalque AC 05.xlsx",     "PR 5.3"),
    (84,  "Placa de Recalque AC 05.xlsx",     "PR 5.4"),
    (85,  "Placa de Recalque AC 05.xlsx",     "PR 5.5"),
    (86,  "Placa de Recalque AC 05.xlsx",     "PR 5.6"),
    (87,  "Placa de Recalque AC 05.xlsx",     "PR 5.7"),
    (88,  "Placa de Recalque AC 05.xlsx",     "PR 5.8"),
    (89,  "Placa de Recalque AC 05.xlsx",     "PR 5.9"),
    (90,  "Placa de Recalque AC 05.xlsx",     "PR 5.10"),
    (91,  "Placa de Recalque AC 05.xlsx",     "PR 5.11"),
    (92,  "Placa de Recalque AC 05.xlsx",     "PR 5.21"),
    (93,  "Placa de Recalque AC 05.xlsx",     "PR 5.22"),
    (94,  "Placa de Recalque AC 05.xlsx",     "PR 5.23"),
    (95,  "Placa de Recalque AC 05.xlsx",     "PR 5.24"),
    (96,  "Placa de Recalque AC 05.xlsx",     "PR 5.25"),
    (97,  "Placa de Recalque AC 05.xlsx",     "PR 5.26"),
    (98,  "Placa de Recalque AC 05.xlsx",     "PR 5.27"),
    (99,  "Placa de Recalque AC 05.xlsx",     "PR 5.28"),
    (100, "Placa de Recalque AC 05.xlsx",     "PR 5.29"),
    (101, "Placa de Recalque AC 05.xlsx",     "PR 5.30"),
    (102, "Placa de Recalque AC 05.xlsx",     "PR 5.41"),
    (103, "Placa de Recalque AC 05.xlsx",     "PR 5.42"),
    (104, "Placa de Recalque AC 05.xlsx",     "PR 5.43"),
    (105, "Placa de Recalque AC 05.xlsx",     "PR 5.44"),
    (106, "Placa de Recalque AC 05.xlsx",     "PR 5.45"),
    (107, "Placa de Recalque AC 05.xlsx",     "PR 5.46"),
    (108, "Placa de Recalque AC 05.xlsx",     "PR 5.47"),
    (109, "Placa de Recalque AC 05.xlsx",     "PR 5.48"),
    (110, "Placa de Recalque AC 05.xlsx",     "PR 5.52"),
    (111, "Placa de Recalque AC 05.xlsx",     "PR 5.53"),
    (112, "Placa de Recalque AC 05.xlsx",     "PR 5.54"),
    (113, "Placa de Recalque AC 05.xlsx",     "PR 5.55"),
    (114, "Placa de Recalque AC 05.xlsx",     "PR 5.56"),
    (115, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.1"),
    (116, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.2"),
    (117, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.3"),
    (118, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.4"),
    (119, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.5"),
    (120, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.6"),
    (121, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.7"),
    (122, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.8"),
    (123, "Placa de Recalque AMPLIAÇÃO.xlsx", "PR A.9"),
    (124, "Placa de Recalque Emergencial.xlsx", "PR 1.1"),
    (125, "Placa de Recalque Emergencial.xlsx", "PR 1.2"),
    (126, "Placa de Recalque Emergencial.xlsx", "PR 1.3"),
    (127, "Placa de Recalque Emergencial.xlsx", "PR 1.4"),
    (128, "Placa de Recalque Emergencial.xlsx", "PR 1.5"),
    (129, "Placa de Recalque Emergencial.xlsx", "PR 2.1"),
    (130, "Placa de Recalque Emergencial.xlsx", "PR 2.2"),
    (131, "Placa de Recalque Emergencial.xlsx", "PR 2.3"),
    (132, "Placa de Recalque Emergencial.xlsx", "PR 2.4"),
    (133, "Placa de Recalque Emergencial.xlsx", "PR 2.5"),
    (134, "Placa de Recalque Emergencial.xlsx", "PR 3.1"),
    (135, "Placa de Recalque Emergencial.xlsx", "PR 3.2"),
    (136, "Placa de Recalque Emergencial.xlsx", "PR 3.3"),
    (137, "Placa de Recalque Emergencial.xlsx", "PR 3.4"),
    (138, "Recalques Celula Pesquisa.xlsx",   "PR01"),
    (139, "Recalques Celula Pesquisa.xlsx",   "PR02"),
    (140, "Recalques Celula Pesquisa.xlsx",   "PR03"),
    (141, "Recalques Celula Pesquisa.xlsx",   "PR04"),
    (142, "Recalques Celula Pesquisa.xlsx",   "PR05"),
    (143, "Recalques Celula Pesquisa.xlsx",   "PR06"),
    (144, "Recalques Celula Pesquisa.xlsx",   "PR07"),
    (145, "Recalques Celula Pesquisa.xlsx",   "PR08"),
    (146, "Recalques Celula Pesquisa.xlsx",   "PR09"),
    (147, "Recalques Celula Pesquisa.xlsx",   "PR10"),
    (148, "Recalques Celula Pesquisa.xlsx",   "PR11"),
    (149, "Recalques Celula Pesquisa.xlsx",   "PR12"),
    (150, "Recalques Celula Pesquisa.xlsx",   "PR13"),
    (151, "Recalques Celula Pesquisa.xlsx",   "PR14"),
    (152, "Recalques Celula Pesquisa.xlsx",   "PR15"),
]

# Nome da tabela onde serão armazenados os resultados (ajuste conforme necessário):
TABELA_RESULTADOS = "verificacao_planilhas"

def criar_tabela_if_not_exists():
    """
    Cria a tabela de resultados no banco de dados se ela ainda não existir.
    """
    # Obter o caminho absoluto do banco de dados
    banco_absoluto = os.path.abspath(BANCO_DADOS)
    conexao = sqlite3.connect(banco_absoluto)
    c = conexao.cursor()
    c.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABELA_RESULTADOS} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_arquivo TEXT,
            nome_planilha TEXT,
            ultima_linha_valida INTEGER,
            data_registro TEXT,
            data_processamento TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conexao.commit()
    conexao.close()


def eh_celula_vazia_ou_zerada(valor):
    """
    Verifica se o valor da célula é vazio, None, string vazia, ou zero.
    """
    if valor is None:
        return True
    
    # Converte tudo para string para simplificar a checagem
    str_val = str(valor).strip()
    
    if str_val == "" or str_val.lower() == "none":
        return True
    
    # Checar se é número e igual a zero
    try:
        val_float = float(str_val.replace('.', '').replace(',', '.'))  # remove separador de milhar se houver, e converte vírgula decimal
        if abs(val_float) < 1e-14:
            return True
    except:
        pass
    
    return False


def padronizar_celula(valor):
    """
    Padroniza o valor para efeito de comparação de duplicidade.
    - Datas -> string no formato YYYY-MM-DD (ou algo coerente)
    - Números -> string normalizada (usar '.' como separador decimal)
    - Vazios/None/Zero -> retorna '0' (como referência).
    - Texto -> texto strip.
    """
    if valor is None:
        return "0"
    
    # Se for data do tipo datetime
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    
    # Tentar identificar se é string contendo data em português:
    # AQUI você pode ajustar a leitura. Exemplo:
    str_val = str(valor).strip()
    
    # Se estiver no Excel como texto "dd/mmm/yy" ou "dd-mmm-yy" em pt-BR e
    # precisar converter manualmente, você pode criar uma rotina extra.
    # Exemplo (bem simplificado, poderia precisar de um dicionário de meses):
    possible_date = try_parse_date_ptbr(str_val)
    if possible_date is not None:
        return possible_date.strftime("%Y-%m-%d")
    
    # Tentar converter para float e ver se é praticamente zero
    try:
        val_float = float(str_val.replace('.', '').replace(',', '.'))
        # Se for zero, padroniza com "0"
        if abs(val_float) < 1e-14:
            return "0"
        else:
            # Mantém com o mesmo número de casas decimais que a string possui (opcional)
            return str_val.replace(',', '.')
    except:
        pass
    
    # Se não for número e não for data, retorna o texto no strip
    if str_val == "":
        return "0"
    return str_val

def try_parse_date_ptbr(texto_data):
    """
    Tenta converter datas em formato pt-BR (ex: "09/set/24" ou "9-mar-2024") em datetime.
    Caso não consiga, retorna None.
    """
    # Você pode ter que ajustar esse mapa se o Excel tiver variações de abreviação.
    # Exemplo: "abr" = "04", "ago" = "08", "set" = "09", "dez" = "12", etc.
    import re
    meses_pt = {
        'jan': '01', 'fev': '02', 'mar': '03', 'abr': '04',
        'mai': '05', 'jun': '06', 'jul': '07', 'ago': '08',
        'set': '09', 'out': '10', 'nov': '11', 'dez': '12'
    }
    # Regex para capturar dia, mes(3 letras) e ano (2 ou 4 dígitos)
    # Ex: 09/set/24 ou 9-mar-2024 etc.
    padrao = r'^(\d{1,2})[/-](\w{3})[/-](\d{2,4})$'
    match = re.match(padrao, texto_data.lower())
    if match:
        dia = match.group(1)
        mes_abrev = match.group(2)
        ano = match.group(3)
        if mes_abrev in meses_pt:
            mes = meses_pt[mes_abrev]
            # Ajustar ano para 20xx ou 19xx, etc. Supondo 20xx:
            if len(ano) == 2:
                # Ex: '24' -> 2024
                ano = f"20{ano}"
            try:
                data_str = f"{ano}-{mes}-{int(dia):02d}"
                return datetime.strptime(data_str, "%Y-%m-%d")
            except:
                return None
    return None


def linha_esta_vazia(ws, row):
    """
    Verifica se a linha 'row' (inteiro) está completamente vazia de A a J.
    """
    for col in range(1, 11):  # colunas 1=A, 2=B, ..., 10=J
        cell = ws.cell(row=row, column=col)
        # Ignorar merged cells "fantasmas"
        if isinstance(cell, MergedCell):
            continue
        
        if not eh_celula_vazia_ou_zerada(cell.value):
            return False
    return True


def ler_data_da_celula(cell):
    """
    Tenta retornar um objeto datetime se a célula realmente for data (ou for
    texto que se converta). Caso não seja data ou não consiga converter, retorna None.
    """
    val = cell.value
    if val is None:
        return None
    
    if isinstance(val, datetime):
        return val  # Já é datetime
    else:
        # Se é string, tenta fazer parse
        return try_parse_date_ptbr(str(val).strip())


def processar_arquivo_e_planilha(nome_arquivo, nome_planilha):
    """
    Abre o arquivo Excel, localiza a planilha, encontra a primeira data de 2024
    na coluna A, processa até 100 linhas após essa data, apaga as vazias e remove
    duplicidades nos últimos 10 registros. Retorna (ultima_linha_valida, data_ultima_linha)
    para gravação em banco.
    """
    caminho = os.path.join(DIRETORIO_ARQUIVOS, nome_arquivo)
    
    # Carrega o workbook com openpyxl:
    wb = load_workbook(filename=caminho, data_only=False, keep_vba=True, keep_links=True)
    
    if nome_planilha not in wb.sheetnames:
        # Se a planilha não existir, avisar e sair
        st.warning(f"Planilha '{nome_planilha}' não encontrada em '{nome_arquivo}' :exclamation:")
        return None, None
    
    ws = wb[nome_planilha]
    
    # 1) Encontrar primeira data >= 01/jan/2024 na coluna A
    primeira_data_2024_row = None
    # Vamos considerar que não sabemos até onde vai a planilha, mas iremos até o "max_row" do sheet.
    max_row = ws.max_row
    
    for row in range(1, max_row + 1):
        cell_data = ler_data_da_celula(ws.cell(row=row, column=1))
        if cell_data is not None:
            # Checar se é >= 2024-01-01
            if cell_data >= datetime(2024, 1, 1):
                primeira_data_2024_row = row
                break
    
    if not primeira_data_2024_row:
        st.info(f"Não há data de 2024 na planilha '{nome_planilha}' em '{nome_arquivo}'.")
        wb.close()
        return None, None
    
    # 2) A partir dessa linha, verificar até 100 linhas abaixo
    #    e remover as que estiverem completamente vazias de A a J.
    linha_inicial = primeira_data_2024_row
    linha_final = min(primeira_data_2024_row + 100, max_row)
    
    # Precisamos checar as linhas de baixo para cima, ao remover dinamicamente
    # (para não alterar a numeração das linhas que ainda não checamos).
    linhas_removidas = 0
    for row in range(linha_final, linha_inicial - 1, -1):
        if linha_esta_vazia(ws, row):
            ws.delete_rows(row, 1)
            linhas_removidas += 1
    
    st.info(f"{linhas_removidas} linha(s) em branco removidas na planilha {nome_planilha} do arquivo {nome_arquivo}.")
    
    # 3) Agora, verificar duplicidades nos últimos 10 registros.
    #    Precisamos encontrar o "final" atual da planilha (pois removemos linhas).
    max_row_atual = ws.max_row
    # Pegar até 10 últimas linhas que tenham dados (pelo menos na coluna A, por exemplo).
    # Vamos varrer de baixo para cima e montar uma lista (até 10)
    ultimas_linhas = []
    contador_coletadas = 0
    row_atual = max_row_atual
    
    while row_atual >= 1 and contador_coletadas < 10:
        # Verifica se a coluna A tem algo (ou se a linha não está inteiramente vazia).
        if not linha_esta_vazia(ws, row_atual):
            ultimas_linhas.append(row_atual)
            contador_coletadas += 1
        row_atual -= 1
    
    # Agora as `ultimas_linhas` estão em ordem decrescente, vamos inverter
    ultimas_linhas.reverse()  # ordem crescente
    if not ultimas_linhas:
        # Não há linhas para analisar
        wb.save(caminho)
        wb.close()
        return None, None
    
    # 4) Remover duplicidades dentro dessas últimas linhas. Se duas linhas tiverem
    #    A-J exatamente iguais (tratando vazios/zeros como iguais, datas etc.).
    # Vamos fazer um set para armazenar "assinaturas" de cada linha.
    # Se aparecer uma assinatura repetida, remove a linha duplicada.
    
    assinaturas = []
    linhas_duplicadas = []
    
    for r in ultimas_linhas:
        # Montar a assinatura do row (colunas A a J)
        row_values = []
        for c in range(1, 11):
            cell_val = ws.cell(row=r, column=c).value
            row_values.append(padronizar_celula(cell_val))
        
        assinatura = tuple(row_values)  # tupla para poder colocar em set
        
        if assinatura in assinaturas:
            # achamos duplicada
            linhas_duplicadas.append(r)
        else:
            assinaturas.append(assinatura)
    
    # Remover as linhas duplicadas (também de baixo para cima)
    for r in sorted(linhas_duplicadas, reverse=True):
        ws.delete_rows(r, 1)
    
    if linhas_duplicadas:
        st.warning(f"{len(linhas_duplicadas)} linha(s) duplicadas removidas na planilha '{nome_planilha}'.")
    else:
        st.success(f"Nenhuma duplicidade encontrada nos últimos 10 registros de '{nome_planilha}'. :thumbsup:")
    
    # 5) Identificar a última linha com informação válida agora:
    max_row_final = ws.max_row
    ultima_linha_valida = None
    data_ultima_linha = None
    
    # Percorre de baixo para cima até achar uma linha não vazia
    for row in range(max_row_final, 0, -1):
        if not linha_esta_vazia(ws, row):
            ultima_linha_valida = row
            # Tentar ler a data (coluna A)
            cell_data = ler_data_da_celula(ws.cell(row=row, column=1))
            if cell_data is not None:
                data_ultima_linha = cell_data.strftime("%Y-%m-%d")
            else:
                data_ultima_linha = str(ws.cell(row=row, column=1).value)
            break
    
    # Salvar as alterações no Excel (preservando formatação existente o máximo possível)
    wb.save(caminho)
    wb.close()
    
    return ultima_linha_valida, data_ultima_linha


def inserir_registro_no_banco(nome_arquivo, nome_planilha, ultima_linha, data_registro):
    """
    Insere no banco de dados sqlite o registro de última linha e data.
    """
    # Obter o caminho absoluto do banco de dados
    banco_absoluto = os.path.abspath(BANCO_DADOS)
    conexao = sqlite3.connect(banco_absoluto)
    c = conexao.cursor()
    c.execute(f"""
        INSERT INTO {TABELA_RESULTADOS} 
        (nome_arquivo, nome_planilha, ultima_linha_valida, data_registro)
        VALUES (?, ?, ?, ?)
    """, (nome_arquivo, nome_planilha, ultima_linha, data_registro))
    conexao.commit()
    conexao.close()


def main():
    st.title("Verificação de Registros em Planilhas Excel")
    if st.button("Iniciar Processamento"):
        st.status("Verificando existência da tabela de resultados...", state="running")
        criar_tabela_if_not_exists()
        st.success("Tabela de resultados pronta (ou já existia).", icon="✅")
        
        total_planilhas = len(PLANILHAS_MAPEAMENTO)
        progresso = st.progress(0, text="Iniciando verificação de planilhas...")
        
        for idx, (id_planilha, nome_arquivo, nome_planilha) in enumerate(PLANILHAS_MAPEAMENTO, start=1):
            with st.spinner(f"Processando arquivo '{nome_arquivo}', planilha '{nome_planilha}'..."):
                st.info(f"Lendo planilha {idx}/{total_planilhas} - [**{nome_planilha}**] do arquivo [**{nome_arquivo}**].")
                
                ultima_linha, data_ult_linha = processar_arquivo_e_planilha(nome_arquivo, nome_planilha)
                
                if ultima_linha is not None:
                    # gravar no banco
                    inserir_registro_no_banco(nome_arquivo, nome_planilha, ultima_linha, data_ult_linha)
                    st.success(f"Atualizado no banco: Última linha válida = {ultima_linha}, Data = {data_ult_linha}.", icon="💾")
                else:
                    st.warning(f"Não foi possível determinar última linha válida em '{nome_planilha}' / '{nome_arquivo}'.")
            
            progresso.progress(idx / total_planilhas, text=f"Progresso: {idx}/{total_planilhas} planilhas")
        
        st.success("Processamento finalizado para todos os arquivos!", icon="🎉")

        # Exemplo de DataFrame para exibir sem erro de Arrow:
        data = {
            "Data": ["2025-02-10", datetime(2025, 2, 12), "texto", 20250101],
            "Coordenadas Norte": [100.0, "200", None, "erro"]
        }
        df = pd.DataFrame(data)
        df = df.astype(str)  # converte todas as colunas para string
        st.dataframe(df)
    else:
        st.info("Clique no botão acima para iniciar o processamento.")


if __name__ == "__main__":
    main()
